#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2022/8/15 18:17
# @Author : ma.fei
# @File : demo.py
# @Software: PyCharm

import sys
import gitlab
import openpyxl
import threading

#获取所有的user
def getAllUsers():
    usersli = []
    client = gitlab.Gitlab(private_host, private_token=private_token)
    users = client.users.list(all=True)
    for user in users:
        usersli.append(user.username)
    return usersli

#获取所有的project
def getAllProjects():
    client = gitlab.Gitlab(private_host, private_token=private_token)
    projects = client.projects.list(all=True)
    return projects

#获取project下所有的branche
def getAllBranchByProject(project):
    try:
        branches = project.branches.list()
        return branches
    except:
        return ""

#获取project和branch下的commit
def getCommitByBranch(project, branch):
    author_commits = []
    commits = project.commits.list(all=True, ref_name=branch.name)
    for commit in commits:
        committer_email = commit.committer_email
        title = commit.title
        message = commit.message
        #if ('Merge' in message) or ('Merge' in title):
        #    print('Merge跳过')
        #    continue
        #else:
        author_commits.append(commit)
    return author_commits

def calculate(data):
    record = {}
    addacount = 0
    deletecount = 0
    totaolcount = 0
    for i in data:
        # print(i)
        addacount += int(i['additions'])
        deletecount += int(i['deletions'])
        totaolcount += int(i['total'])
    record['additions'] = addacount
    record['deletions'] = deletecount
    record['total'] = totaolcount
    return record

#获取project项目下commit对应的code
def getCodeByCommit(commit, project):
    commit_info = project.commits.get(commit.id)
    code = commit_info.stats
    return code

def getAuthorCode(project,fenzhi):
    # print("project:%s" % project)
    users = getAllUsers()
    branches = getAllBranchByProject(project)
    if branches == "":
        pass
    else:
        for branch in branches:
            # print("branch#####",branch.name)
            if branch.name == fenzhi:
                #print("branch:%s" % branch)
                #print('获取工程', project.name, '分支', branch.name, "的提交记录")
                branchdata = {}
                branchdata['group'] = project.name_with_namespace.split("/")[0]
                branchdata['projectname'] = project.name
                branchdata['branchename'] = branch.name
                author_commits = getCommitByBranch(project, branch)
                # print(author_commits)
                codes = []
                res1 = []
                for commit in author_commits:
                    #print('获取提交', commit.id, "的代码量")
                    code = getCodeByCommit(commit, project)
                    # print(commit,code)
                    # print(code)
                    # print(commit)
                    # print(commit.committer_name)
                    codes.append(code)
                    # for user in users:
                    #     if commit.committer_name == user:
                    #         res1.append(commit)
                record = calculate(codes)
                branchdata['commitcount'] = len(author_commits)
                branchdata['codecount'] = record
                data.append(branchdata)
    # print(codes)
    # print(calculate(codes))
    # print(data)
    # for res in res1:
    #     print(res)
    return data


# client = gitlab.Gitlab('https://gitlab.sheyushop.com', private_token='dP4WW4afRfa-JyscunJJ', timeout=5, api_version='4')
# client.auth()
# project = client.projects.list()
# for pro in project:
#     #print(pro)
#     commits = pro.commits.list(since='2000-01-01T00:00:00Z', ref_name='dev')
#     for c in commits:
#         print(pro.commits.get(c.id))


if __name__ == '__main__':
    private_token = 'dP4WW4afRfa-JyscunJJ'
    private_host = 'https://gitlab.sheyushop.com'

    data = []
    thread_list = []
    projects = getAllProjects()
    i_counter = 0
    for i in projects:
        print('project=',i)
        i_counter=i_counter+1
        branches = getAllBranchByProject(i)
        for j in branches:
            #getAuthorCode(i, j.name)
            t = threading.Thread(target=getAuthorCode, args=(i, j.name))
            thread_list.append(t)

        # if i_counter % 10 ==0:
        #    break

    for threadname in thread_list: threadname.start()
    for threadname in thread_list: threadname.join()

    for i in data:
        print(i)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0,title='demo')
    file ='exp_bbtj_gitlab.xlsx'

    r = 1
    c = 1

    # write header
    ws.cell(column=c, row=r, value='项目组名')
    ws.cell(column=c+1, row=r, value='项目名称')
    ws.cell(column=c+2, row=r, value='分支名称')
    ws.cell(column=c+3, row=r, value='提交次数')
    ws.cell(column=c+4, row=r, value='代码行数')

    # write body
    r = r+1
    for i in data:
       ws.cell(row=r, column=c, value=i['group'])
       ws.cell(row=r, column=c+1, value=i['projectname'])
       ws.cell(row=r, column=c+2, value=i['branchename'])
       ws.cell(row=r, column=c+3, value=i['commitcount'])
       ws.cell(row=r, column=c+4, value=i['codecount']['total'])
       r =r +1

    wb.save(file)
