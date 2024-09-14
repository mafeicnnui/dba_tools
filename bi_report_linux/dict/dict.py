#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os,sys
import pymysql as py
import pandas as pd
import openpyxl
import warnings
from openpyxl.styles import Side,Border,PatternFill,Font,Alignment

def write_dict():
    conn = py.connect(host='10.2.39.201', user='root',passwd='root123HOPSON', port=4000,database='test')
    db_names = """SELECT table_schema AS schema_name ,COUNT(0) AS table_count
                   FROM information_schema.`tables` 
                  WHERE table_schema NOT IN(
                    'mysql', 'sys', 'INFORMATION_SCHEMA', 'PERFORMANCE_SCHEMA', 'METRICS_SCHEMA', 'INSPECTION_SCHEMA',
                    'zabbix','apolloconfigdb_v2','apolloportaldb_v2','puppet','xx1-job',
                    'asset','hft_business_pay','hft_commission',
                    'xxl-job','test','mq_cloud',
                    'business_finance','ke','lightning_task_info','nacos_config') 
                     GROUP BY 1 ORDER BY 1 """
    tab_names = """SELECT table_name,table_comment
                    FROM information_schema.tables
                    WHERE table_schema='{}' 
                    -- and table_name in('asset_area','asset_build') 
                     ORDER BY table_schema ,table_name"""
    col_names = """SELECT 
                      c.ordinal_position,
                      c.column_name AS '字段名',
                      c.data_type AS '类型',
                      IF(c.is_nullable='YES','否','是') AS '是否非空',
                      IF(c.column_key='PRI','是','否') AS '是否主键',
                      c.column_comment AS '注释'
                    FROM information_schema.columns c ,information_schema.tables t
                    WHERE c.table_schema=t.table_schema AND c.table_name=t.table_name
                      AND c.table_schema='{}' AND c.table_name='{}'
                     ORDER BY c.table_name,c.ordinal_position"""
    df_db_names = pd.read_sql(db_names,con=conn,index_col='schema_name')

    dict_wb = openpyxl.Workbook('./dict.xlsx')
    dict_wb.create_sheet('db_lists')
    for db, _ in df_db_names.iterrows():
        dict_wb.create_sheet(db)
    dict_wb.save('./dict.xlsx')

    dict_wb = openpyxl.load_workbook('./dict.xlsx')
    line = Side(style="thin", color="000000")
    border = Border(left=line, right=line, top=line, bottom=line)
    color = PatternFill(patternType="solid", start_color='D9E1F2')
    folded_style = Alignment(wrap_text=True)
    r = 1
    sheet = dict_wb['db_lists']
    sheet.cell(r, 1).value = '库名'
    sheet.cell(r, 2).value = '表数量'
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 10
    for row in sheet['A1':'B1']:
        for c in row:
            c.border = border
            c.fill = color
    r += 1
    start_row = r
    for db, row in df_db_names.iterrows():
        sheet.cell(r, 1).value = db
        sheet.cell(r, 2).value = row['table_count']
        r += 1
    end_row = r - 1
    for row in sheet['A' + str(start_row):'B' + str(end_row)]:
        for c in row:
            c.border = border
    # dict_wb.save('./out/dict.xlsx')
    # sys.exit(0)

    for db,_ in df_db_names.iterrows():
        sheet = dict_wb[db]
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 10
        sheet.column_dimensions["F"].width = 30
        r = 1
        df_tab_names = pd.read_sql(tab_names.format(db), con=conn, index_col='table_name')
        for tab, tab_rows in df_tab_names.iterrows():
            sheet.cell(r, 1).value = '表名称 : {}（{}）'.format(tab,tab_rows['table_comment'])
            sheet.cell(r + 1, 1).value = '编号'
            sheet.cell(r + 1, 2).value = '字段名'
            sheet.cell(r + 1, 3).value = '类型'
            sheet.cell(r + 1, 4).value = '是否非空'
            sheet.cell(r + 1, 5).value = '是否主键'
            sheet.cell(r + 1, 6).value = '注释'
            for row in sheet['A'+str(r + 1):'F'+str(r + 1)]:
                for c in row:
                    c.border = border
                    c.fill = color

            df_col_names = pd.read_sql(col_names.format(db,tab), con=conn, index_col='ordinal_position')
            r += 2
            start_row = r
            for key, value in df_col_names.iterrows():
                sheet.cell(r, 1).value = key
                sheet.cell(r, 2).value = value['字段名']
                sheet.cell(r, 3).value = value['类型']
                sheet.cell(r, 4).value = value['是否非空']
                sheet.cell(r, 5 ).value = value['是否主键']
                sheet.cell(r, 6).value = value['注释']
                sheet.cell(r, 6).alignment = folded_style
                r += 1
            end_row = r - 1
            for row in sheet['A' + str(start_row):'F' + str(end_row)]:
                for c in row:
                    c.border = border
            r += 2
    dict_wb.save('./out/dict.xlsx')

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    write_dict()