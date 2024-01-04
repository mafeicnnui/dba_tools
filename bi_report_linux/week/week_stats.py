#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/28 9:26
# @Author : ma.fei
# @File : sales_points.py.py
# @Software: PyCharm

# 各项目销售积分&运营积分(周报)
'''
 sudo pip3 install pandas==1.1.5 -i https://pypi.douban.com/simple
 sudo pip3 install pymysql -i https://pypi.douban.com/simple
 sudo pip3 install openpyxl -i https://pypi.douban.com/simple
'''
#导入模块
import pymysql as py
import pandas as pd
import datetime as dt
import os
import zipfile
import pathlib
import smtplib
import openpyxl
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime,timedelta

def zip_extract_to(path_zip, path_aim):
    '''
     https://cloud.tencent.com/developer/article/1908473
     作用：解压压缩包
     参数一：压缩包位置
     参数二：解压后的路径

     # 要解压的zip文件路径
     path_zip = r'D:\lizhi\压缩测试\chromedriver_win32.zip'
     # 要解压到的位置
     path_aim = r'D:\lizhi\压缩测试'
     # 解压压缩包
     zip_extract_to(path_zip, path_aim)
    '''
    z = zipfile.ZipFile(path_zip, 'r')
    for p in z.namelist():
        z.extract(p, path_aim)
    z.close()

def file_to_zip(path_original, path_zip):
    '''
     作用：压缩文件到指定压缩包里
     参数一：压缩文件的位置
     参数二：压缩后的压缩包
    '''
    # 提前读取，避免把压缩包自己加上去
    # 这里用list()做一个克隆提前执行下，不然会在后面循环时才执行这一引用，如果压缩包在这个路径下，会将它读取进来。
    f_list = list(pathlib.Path(path_original).glob("**/*"))
    z = zipfile.ZipFile(path_zip, 'w')
    #print('len(path_original)=',len(path_original))
    for f in f_list:
        #print('f=',f,str(f),str(f)[len(path_original):])
        z.write(f, str(f)[len(path_original):])
    z.close()

def get_last_month_first_last_day():
    now = datetime.now()
    this_month_start = datetime(now.year, now.month, 1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    v_last_month_start = datetime.strftime(last_month_start, "%Y-%m-%d")
    v_last_month_end = datetime.strftime(last_month_end, "%Y-%m-%d")
    return v_last_month_start, v_last_month_end

def get_monday_to_sunday(today, weekly=0):
    """
    :function: 获取指定日期的周一和周日的日期
    :param today: '2021-11-16'; 当前日期：today = datetime.now().strftime('%Y-%m-%d')
    :param weekly: 获取指定日期的上几周或者下几周，weekly=0当前周，weekly=-1上一周，weekly=1下一周
    :return: 返回指定日期的周一和周日日期
    :return_type: tuple
    """
    last = weekly * 7
    today = datetime.strptime(str(today), "%Y-%m-%d")
    monday = datetime.strftime(today - timedelta(today.weekday() - last), "%Y-%m-%d")
    monday_ = datetime.strptime(monday, "%Y-%m-%d")
    sunday = datetime.strftime(monday_ + timedelta(monday_.weekday() + 6), "%Y-%m-%d")
    return monday, sunday

class SendMail(object):
    '''
    https://cloud.tencent.com/developer/article/1598257
    '''
    def __init__(self,sender,Cc,title,content):
        self.sender = sender.split(',')  #发送地址
        self.Cc = Cc.split(',')  # 抄送地址
        self.title = title  # 标题
        self.content = content  # 发送内容
        self.sys_sender = '190343@lifeat.cn'  # 系统账户
        self.sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码

    def send(self,file_list):
        """
        发送邮件
        :param file_list: 附件文件列表
        :return: bool
        """
        try:
            # 创建一个带附件的实例
            msg = MIMEMultipart()
            # 发件人格式
            msg['From'] = formataddr(["技术服务部", self.sys_sender])
            # 收件人格式
            print('a=',[formataddr(["", u]) for u in self.sender])
            msg['To'] = sender

            # 抄送人
            msg['Cc'] = Cc

            # 邮件主题
            msg['Subject'] = self.title

            # 邮件正文内容
            msg.attach(MIMEText(self.content, 'plain', 'utf-8'))

            # 多个附件
            for file_name in file_list:
                print("file_name",file_name)
                # 构造附件
                xlsxpart = MIMEApplication(open(file_name, 'rb').read())
                # filename表示邮件中显示的附件名
                xlsxpart.add_header('Content-Disposition','attachment',filename = '%s'%file_name)
                msg.attach(xlsxpart)

            # SMTP服务器
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465,timeout=10)
            # 登录账户
            server.login(self.sys_sender, self.sys_pwd)
            # 发送邮件
            #server.sendmail(self.sys_sender, [self.sender, ], msg.as_string())
            server.sendmail(self.sys_sender,
                            [formataddr(["", u]) for u in self.sender]+
                            [formataddr(["", u]) for u in self.Cc],
                            msg.as_string())
            # 退出账户
            server.quit()
            return True
        except Exception as e:
            print(e)
            return False

def stats_sales_points(p_rq_start,p_rq_end,p_week_rq):
    开始日期, 结束日期,周报日期 = p_rq_start,p_rq_end,p_week_rq
    date1=dt.datetime.strptime(开始日期,'%Y-%m-%d').date()
    date2=dt.datetime.strptime(结束日期,'%Y-%m-%d').date()

    #计时
    销售积分运行开始时间=dt.datetime.now()
    #连接数据库，执行SQL语句
    连接对象=py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com',user='hopsonone_do',passwd='8Loxk2IJxaenJkE3',database='hopsonone_do')
    #销售积分SQL语句
    销售积分sql="""
    select 
    d1.market_name as '项目名称',
    d1.store_id as '商家编码',
    d1.`商家名称`,
    d1.`铺位号`,
    d1.`楼层`,
    d1.`状态`,
    d1.`商家类型`,
    d2.`业态`,
    a.`商家销售额(元)`,
    a.`商家交易笔数(笔)`,
    ifnull(b1.point_value,0) as '积分规则(1元=积分数量)',
    ifnull(round(a.`商家销售额(元)`* ifnull(b1.point_value,0),2),0) as '应积分数量',
    round(ifnull(a1.`会员积分数量`,0),2) as '会员积分数量',
    concat(round(ifnull(a1.`会员积分数量`/(a.`商家销售额(元)` * ifnull(b1.point_value,0)),0)*100,2),'%%') as '会员积分数量占比',
    ifnull(a1.`会员积分笔数`,0) as '会员积分笔数',
    concat(round(ifnull(a1.`会员积分笔数`/a.`商家交易笔数(笔)`,0)*100,2),'%%') as '会员积分笔数占比',
    round(ifnull(a1.`总消费金额(元)`,0),2) as '总消费金额(元)',
    round(ifnull(a1.`微信积分数量`,0)+ifnull(a1.`支付宝积分数量`,0),2) as '支付即积分总数量',
    concat(round(ifnull((a1.`微信积分数量`+a1.`支付宝积分数量`)/a1.`会员积分数量`,0)*100,2),'%%') as '支付即积分数量占比',
    round(ifnull(a1.`微信积分笔数`,0)+ifnull(a1.`支付宝积分笔数`,0),2) as '支付即积分总笔数',
    concat(round(ifnull((a1.`微信积分笔数`+a1.`支付宝积分笔数`)/a1.`会员积分笔数`,0)*100,2),'%%') as '支付即积分笔数占比',
    ifnull(a1.`微信积分数量`,0) as '微信积分数量',
    ifnull(a1.`微信积分笔数`,0) as '微信积分笔数',
    round(ifnull(a1.`微信消费金额(元)`,0),2) as '微信消费金额(元)',
    ifnull(a1.`支付宝积分数量`,0) as '支付宝积分数量',
    ifnull(a1.`支付宝积分笔数`,0) as '支付宝积分笔数',
    round(ifnull(a1.`支付宝消费金额(元)`,0),2) as '支付宝消费金额(元)',
    ifnull(a1.`扫码积分数量`,0) as '扫码积分数量',
    ifnull(a1.`扫码积分笔数`,0) as '扫码积分笔数',
    round(ifnull(a1.`扫码积分消费金额(元)`,0),2) as '扫码积分消费金额(元)',
    ifnull(a1.`客服积分调整数量`,0) as '客服积分调整数量',
    ifnull(a1.`客服积分调整笔数`,0) as '客服积分调整笔数',
    round(ifnull(a1.`客服积分调整消费金额(元)`,0),2) as '客服积分调整消费金额(元)',
    0 as '客服消费积分调整数量',
    0 as '客服消费积分调整笔数',
    ifnull(a1.`当面付积分数量`,0) as '当面付积分数量',
    ifnull(a1.`当面付积分笔数`,0) as '当面付积分笔数',
    round(ifnull(a1.`当面付消费金额(元)`,0),2) as '当面付消费金额(元)'
    from 
    (select ma.market_name,
    d.store_id,
    d.store_name as '商家名称',
    d.store_berth as '铺位号',
    s.dic_desc as '楼层',
    (case d.apply_status when 1 then '有效' when 0 then '失效' end) as '状态',
    (case d.store_category when 1 then '正式商户' when 2 then '虚拟商户' end) as'商家类型'
    from merchant_entity.entity_store d,
    hopsonone_cms.sys_dic s,
    merchant_entity.market ma
    where d.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    and d.market_id=s.market_id
    and d.floor_type=s.dic_value
    and s.type_name='floortype'
    and d.market_id=ma.id
    GROUP BY ma.market_name,d.store_id) d1
    left join (select ma.market_name,
    d.store_id,
    s.dic_desc as '业态'
    from  merchant_entity.entity_store d,
    hopsonone_cms.sys_dic s,
    merchant_entity.market ma
    where d.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    and d.market_id=s.market_id
    and d.format_type=s.dic_value
    and s.type_name='operationtype'
    and d.market_id=ma.id
    GROUP BY ma.market_name,d.store_id) d2
    on d1.market_name=d2.market_name and d1.store_id=d2.store_id
    
    left join (select ma.market_name,c.business_id,b.point_value
    from hopsonone_point.point_rule b,
    hopsonone_point.point_business_rule c,
    merchant_entity.entity_store bu,
    merchant_entity.market ma
    where c.rule_id=b.id
    and b.status=0
    and bu.store_id = c.business_id
    and b.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    and b.market_id=ma.id 
    GROUP BY ma.market_name,c.business_id) b1
    on d1.market_name=b1.market_name and d1.store_id=b1.business_id
    
    left join 
    (select ma.market_name,b.store_id,
    ifnull(round(sum(a.adjust_tax_sales_amount_month)/100,2),0) as '商家销售额(元)',
    ifnull(sum(a.adjust_trade_count_day),0) as '商家交易笔数(笔)'
    from 
    (select a.business_id,a.adjust_tax_sales_amount_month,a.adjust_trade_count_day
    from shop_side_operation_real_time.sales_report_details a
    where a.trade_date >= %(date_op)s
    and a.trade_date<=%(date_ed)s) a,
    merchant_entity.entity_store b,
    merchant_entity.market ma
    where a.business_id = b.store_id
    and b.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    and b.market_id=ma.id
    group by ma.market_name,b.store_id) a
    on a.market_name=d1.market_name and a.store_id=d1.store_id
    
    left join(select 
    ma.market_name,
    a.business_id,
    sum(case when a.state = 0 then ifnull(a.point,0) 
    when a.state = 1 then -ifnull(a.point,0) 
    end ) as '会员积分数量',
    sum(case when a.state = 0 then 1 
    when a.state = 1 then -1 
    end ) as '会员积分笔数',
    sum(case when a.state = 0 then ifnull(a.consume_amount,0) 
    when a.state = 1 then -ifnull(a.consume_amount,0) 
    end ) as '总消费金额(元)',
    sum(case when a.behavior_name='微信' then a.point else 0 end) as '微信积分数量',
    sum(case when a.behavior_name='微信' then 1 else 0 end) as '微信积分笔数',
    sum(case when a.behavior_name='微信' then a.consume_amount else 0 end) as '微信消费金额(元)',
    sum(case when a.behavior_name='支付宝' then a.point else 0 end) as '支付宝积分数量',
    sum(case when a.behavior_name='支付宝' then 1 else 0 end) as '支付宝积分笔数',
    sum(case when a.behavior_name='支付宝' then a.consume_amount else 0 end) as '支付宝消费金额(元)',
    sum(case when a.behavior_name='消费-小票' then a.point else 0 end) as '扫码积分数量',
    sum(case when a.behavior_name='消费-小票' then 1 else 0 end) as '扫码积分笔数',
    sum(case when a.behavior_name='消费-小票' then a.consume_amount else 0 end) as '扫码积分消费金额(元)',
    sum(case when a.behavior_name='客服积分调整' and a.state = 0 then a.point 
    when a.behavior_name='客服积分调整' and a.state = 1 then -ifnull(a.point,0) 
    else 0 end) as '客服积分调整数量',
    sum(case when a.behavior_name='客服积分调整' and a.state = 0 then 1 
    when a.behavior_name='客服积分调整' and a.state = 1 then -1
    else 0 end) as '客服积分调整笔数',
    sum(case when a.behavior_name='客服积分调整' and a.state = 0 then a.consume_amount
    when a.behavior_name='客服积分调整' and a.state = 1 then -ifnull(a.consume_amount,0) 
    else 0 end) as '客服积分调整消费金额(元)',
    sum(case when a.behavior_name='当面付' then a.point else 0 end) as '当面付积分数量',
    sum(case when a.behavior_name='当面付' then 1 else 0 end) as '当面付积分笔数',
    sum(case when a.behavior_name='当面付' then a.consume_amount else 0 end) as '当面付消费金额(元)' 
    from 
    (select a.market_id,a.business_id,a.point,a.consume_amount/100 as consume_amount,a.behavior_name,a.state
    from hopsonone_point_real_time.members_points_detail a
    where a.consume_time >=%(date_op)s
    and a.consume_time<=concat(%(date_ed)s," 23:59:59")) a,
    merchant_entity.entity_store b,
    merchant_entity.market ma
    where a.business_id=b.store_id
    and a.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    and a.market_id=ma.id
    group by a.business_id) a1
    on d1.market_name=a1.market_name and d1.store_id=a1.business_id
    WHERE a.`商家销售额(元)`is not null
    union all
    select ma.market_name as '项目名称',
    '' as '商家编码',
    '客服台' as '商家名称',
    '' as '铺位号',
    '' as '楼层',
    '' as '状态',
    '' as '商家类型',
    '' as '业态',
    '' as '商家销售额(元)',
    '' as '商家交易笔数(笔)',
    '' as '积分规则(1元=积分数量)',
    0 as '应积分数量',
    0 as '会员积分数量',
    0 as '会员积分数量占比',
    0 as '会员积分笔数',
    0 as '会员积分笔数占比',
    0 as '总消费金额(元)',
    0 as '支付即积分总数量',
    0 as '支付即积分数量占比',
    0 as '支付即积分总笔数',
    0 as '支付即积分笔数占比',
    0 as '微信积分数量',
    0 as '微信积分笔数',
    0 as '微信消费金额(元)',
    0 as '支付宝积分数量',
    0 as '支付宝积分笔数',
    0 as '支付宝消费金额(元)',
    0 as '扫码积分数量',
    0 as '扫码积分笔数',
    0 as '扫码积分消费金额(元)',
    0 as '客服积分调整数量',
    0 as '客服积分调整笔数',
    0 as '客服积分调整消费金额(元)',
    pa.`客服消费积分调整数量`,
    pa.`客服消费积分调整笔数`,
    0 as '当面付积分数量',
    0 as '当面付积分笔数',
    0 as '当面付消费金额(元)'
    from merchant_entity.market ma,
    (select a.market_id,sum(a.point) as '客服消费积分调整数量',count(*) as '客服消费积分调整笔数'
    from hopsonone_point_real_time.members_points_detail a
    where a.create_time >=%(date_op)s
    and a.create_time<=concat(%(date_ed)s," 23:59:59")
    and a.behavior=20
    and a.consume_time is null
    and a.market_id in (218,108,237,110,132,213,164,234,278,287,306)
    group by a.market_id) pa
    where ma.id=pa.market_id
    """
    销售积分数据=pd.read_sql(销售积分sql,con=连接对象,params={'date_op':开始日期,'date_ed':结束日期})
    销售积分运行结束时间=dt.datetime.now()
    销售积分运行sql用时=(销售积分运行结束时间-销售积分运行开始时间).seconds
    print(f'销售积分运行SQL时间为{销售积分运行sql用时}秒')
    # 销售积分拆分列=list(销售积分数据['项目名称'].drop_duplicates())
    销售积分拆分列=['成都温江合生汇','上海五角场合生汇','广州海珠合生广场(南)','北京木樨园合生广场','北京合生麒麟新天地','北京朝阳合生汇','上海青浦合生新天地','广州增城合生汇','广州海珠合生新天地','西安南门合生汇','上海静安MOHO']
    for i in 销售积分拆分列:
        销售积分新数据=销售积分数据[销售积分数据['项目名称']==i]
        if (date2 - date1).days > 10:
            项目文件夹 = output_dir+ i +'月报'+ 周报日期
            os.mkdir(项目文件夹)
            销售积分新数据.to_excel(项目文件夹 + '/' + i + '销售积分' + 周报日期 + '.xlsx', index=False)
        else:
            项目文件夹 = output_dir + i +'周报'+ 周报日期
            os.mkdir(项目文件夹)
            销售积分新数据.to_excel(项目文件夹 + '/' + i + '销售积分' + 周报日期 + '.xlsx', index=False)
    #快速积分
    快速积分运行开始时间=dt.datetime.now()
    #快速积分SQL语句
    快速积分sql="""
    # 微信开通店铺算法
    SELECT
    a.market_name as '项目名称',
    a.mid as '会员id',
    d.behavior_name as '渠道',
    date(a.create_time) as '开通时间',
    d.business_id as '商家编码',
    b.business_name as '商家名称',
    b.business_berth as '铺位号',
    d.create_time as '积分时间',
    d.point as '积分数量',
    d.consume_amount as '消费金额'
    FROM
    (SELECT ma.market_name,w.mid,w.create_time
    FROM hopsonone_personal.wx_auth_log w,
    merchant_entity.market ma
    WHERE w.create_time>=%(date_op)s
    and w.create_time<=concat(%(date_ed)s,' 23:59:59')
    and ma.id in (218,108,237,110,132,213,164,234,278,287,306)
    and locate(ma.id,w.market_id)) a,
    (select d.create_time,d.m_id,d.behavior_name,d.business_id,d.point,d.consume_amount/100 as consume_amount
    from hopsonone_point_real_time.members_points_detail d
    where d.state = 0
    and d.behavior_name = '微信') d,
    merchant_entity.business b
    WHERE a.mid = d.m_id 
    and d.create_time>=%(date_op)s
    and d.create_time<=concat(%(date_ed)s,' 23:59:59')
    and date(a.create_time) = date(d.create_time) 
    and d.business_id = b.business_id
    group by a.mid
    union all
    SELECT
    a.market_name as '项目名称',
    a.mid as '会员id',
    d.behavior_name as '渠道',
    date(a.create_time) as '开通时间',
    d.business_id as '商家编码',
    b.business_name as '商家名称',
    b.business_berth as '铺位号',
    d.create_time as '积分时间',
    d.point as '积分数量',
    d.consume_amount as '消费金额'
    FROM
    (SELECT l.mid,m.market_id,l.create_time,ma.market_name
    FROM hopsonone_personal.ali_card_log l,
    hopsonone_members.members m,
    merchant_entity.market ma
    WHERE l.mid = m.m_id
    and ma.id=m.market_id
    and ma.id in (218,108,237,110,132,213,164,234,278,287,306)
    and l.create_time>=%(date_op)s
    and l.create_time<=concat(%(date_ed)s,' 23:59:59')) a,
    (select d.create_time,d.m_id,d.behavior_name,d.business_id,d.point,d.consume_amount/100 as consume_amount
    from hopsonone_point_real_time.members_points_detail d
    where d.state = 0
    and d.behavior_name = '支付宝') d,
    merchant_entity.business b
    WHERE a.mid = d.m_id 
    and d.create_time>=%(date_op)s
    and d.create_time<=concat(%(date_ed)s,' 23:59:59')
    and date(a.create_time) = date(d.create_time) 
    and d.business_id = b.business_id
    GROUP BY a.mid
    """
    快速积分数据=pd.read_sql(快速积分sql,con=连接对象,params={'date_op':开始日期,'date_ed':结束日期})
    快速积分运行结束时间=dt.datetime.now()
    快速积分运行sql用时=(快速积分运行结束时间-快速积分运行开始时间).seconds
    print(f'快速积分运行SQL时间为{快速积分运行sql用时}秒')
    快速积分拆分列=list(快速积分数据['项目名称'].drop_duplicates())
    for i in 快速积分拆分列:
        快速积分新数据=快速积分数据[快速积分数据['项目名称']==i]
        if (date2 - date1).days > 10:
            快速积分新数据.to_excel(output_dir + i+'月报'+周报日期 +'/'+i+'快速积分'+周报日期+'.xlsx',index=False)
        else:
            快速积分新数据.to_excel(output_dir + i + '周报' + 周报日期 + '/' + i + '快速积分' + 周报日期 + '.xlsx', index=False)

    #会员消费SQL语句
    会员消费sql="""
    select date(p.consume_time) as '消费日期',
    date_format(p.consume_time,'%%h:%%m:%%s') as '消费时间',
    concat(" ",p.m_id) as '会员ID',
    case m.sex
    when 1 then '男'
    when 2 then '女'
    when 0 then null
    end as '性别',
    2022-year(m.birthday) as '年龄',
    s.dic_desc as '注册来源',
    c.plate_number as '车牌号',
    m.m_mobile as '手机号',
    p.business_id as '商家编码',
    b.`商家名称`, 
    b.`铺位号`,
    b.`楼层`,
    b.`业态`,
    p.consume_amount as '消费金额（元）',
    p.point as '会员积分数量'
    from (select p.m_id,p.consume_time,p.business_id,
    case when p.state = 0 then ifnull(p.consume_amount/100,0) 
    when p.state = 1 then -ifnull(p.consume_amount/100,0) 
    end  as 'consume_amount',
    case when p.state = 0 then ifnull(p.point,0) 
    when p.state = 1 then -ifnull(p.point,0) 
    end  as 'point'
    FROM hopsonone_point_real_time.members_points_detail p
    where p.market_id=218
    and p.behavior in (8,20,24,26,27,28)
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59')) p
    left join 
    (select m.m_id,m.m_mobile,f.birthday,f.sex,m.m_origin,p.state
    from hopsonone_point_real_time.members_points_detail p,
    hopsonone_members.members m,
    hopsonone_members.members_info f
    where p.market_id=218
    and p.behavior in (8,20,24,26,27,28)
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59')
    and m.m_id=p.m_id
    and m.m_id=f.m_id
    group by m.m_id) m
    on p.m_id=m.m_id
    left join 
    (select distinct s.dic_value,s.dic_desc 
    from hopsonone_cms.system_sys_dic s
    where s.type_name='REGISTER_TYPE') s 
    on m.m_origin=s.dic_value
    left join 
    (SELECT a.m_id,group_concat(a.plate_number separator ',') as 'plate_number'
    from hopsonone_members.members_plate_number a,
    hopsonone_point_real_time.members_points_detail p
    where a.market_id=218
    and a.status=0
    and p.state = 0 
    and p.behavior in (8,20,24,26,27,28)
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59')
    and p.m_id=a.m_id
    GROUP BY a.m_id) c 
    on p.m_id=c.m_id
    left join 
    (select d.market_id,d.business_id,d.business_name as '商家名称',
    d.business_berth as '铺位号',d1.dic_desc as '楼层',d2.dic_desc as '业态'
    from merchant_entity.business d
    left join 
    (select s.market_id,s.dic_value,s.dic_desc
    from hopsonone_cms.sys_dic s
    where s.market_id=218
    and s.type_name='floortype') d1
    on d.market_id=d1.market_id and d.business_fool=d1.dic_value
    left join 
    (select s.market_id,s.dic_value,s.dic_desc
    from hopsonone_cms.sys_dic s
    where s.market_id=218
    and s.type_name='operationtype') d2
    on d.market_id=d2.market_id and d.business_type=d2.dic_value
    where d.market_id=218
    group by d.business_id) b 
    on p.business_id=b.business_id
    order by p.consume_time
    """
    if (date2-date1).days>10:
        会员消费运行开始时间=dt.datetime.now()
        会员消费数据=pd.read_sql(会员消费sql,con=连接对象,params={'date_op':开始日期,'date_ed':结束日期})
        会员消费运行结束时间=dt.datetime.now()
        会员消费运行sql用时=(会员消费运行结束时间-会员消费运行开始时间).seconds
        print(f'会员消费运行SQL时间为{会员消费运行sql用时}秒')
        会员消费数据.to_excel(output_dir + '北京朝阳合生汇月报' + 周报日期 + '/' + '北京朝阳合生汇会员消费' + 周报日期 + '.xlsx', index=False)
    print('文件生成完成!')

def stats_week_report(p_rq_start,p_rq_end,p_week_rq):
    开始日期, 结束日期, 周报日期 = p_rq_start, p_rq_end, p_week_rq
    # 连接数据库，执行SQL语句
    连接对象 = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com',
                        user='hopsonone_do',
                        passwd='8Loxk2IJxaenJkE3',
                        database='hopsonone_do')
    # 创建项目名称ID字典
    market = {'成都温江合生汇': 108, '上海五角场合生汇': 110, '广州海珠合生广场(南)': 132, '北京木樨园合生广场': 164, '北京合生麒麟新天地': 213, '北京朝阳合生汇': 218,
              '上海青浦合生新天地': 234, '广州增城合生汇': 237, '广州海珠合生新天地': 278, '西安南门合生汇': 287, '上海静安MOHO': 306}
    # 遍历字典键值对
    for 项目名称, 项目ID in market.items():
        # 会员注册量SQL语句
        会员拉新sql = """
        select a.market_id,
        count(distinct (case when a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.m_id end)) as '本周',
        count(distinct (case when a.create_time>=DATE_SUB(%(date_op)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.m_id end)) as '较前一周',
        count(distinct (case when a.create_time>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s, INTERVAL 1 year),' 23:59:59') then a.m_id end)) as '同期',
        count(distinct (case when a.create_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.m_id end)) as '本月'
        from hopsonone_members.members a
        where a.market_id=%(market_id)s
        and a.m_status=1
        and a.create_time>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY a.market_id
        """
        # 计时
        会员拉新运行开始时间 = dt.datetime.now()
        会员拉新 = pd.read_sql(会员拉新sql,
                             con=连接对象,
                             params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        会员拉新运行结束时间 = dt.datetime.now()
        会员拉新运行sql用时 = (会员拉新运行结束时间 - 会员拉新运行开始时间).seconds
        print(f'会员拉新运行SQL时间为{会员拉新运行sql用时}秒')

        # 付费会员新增SQL语句
        付费会员新增sql = """
        select o.market_id,
        count((case when o.payment_dt>=%(date_op)s
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.buyer_id end)) as '本周',
        count((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and o.payment_dt<=concat(DATE_SUB(%(date_ed)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then o.buyer_id end)) as '较前一周',
        count((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and o.payment_dt<=concat(DATE_SUB(%(date_ed)s, INTERVAL 1 year),' 23:59:59') then o.buyer_id end)) as '同期',
        count((case when o.payment_dt>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.buyer_id end)) as '本月'
        from hopsonone_order_center.orders_detail_plusvip a,
        hopsonone_order_center.orders o
        where o.market_id=%(market_id)s
        and a.trade_no = o.trade_no
        and o.state in (3, 4, 5)
        and o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY o.market_id
        """
        # 计时
        付费会员新增运行开始时间 = dt.datetime.now()
        付费会员新增 = pd.read_sql(付费会员新增sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        付费会员新增运行结束时间 = dt.datetime.now()
        付费会员新增运行sql用时 = (付费会员新增运行结束时间 - 付费会员新增运行开始时间).seconds
        print(f'付费会员新增运行SQL时间为{付费会员新增运行sql用时}秒')

        # 付费会员收益SQL语句
        付费会员收益sql = """
        SELECT o.market_id,
        sum((case when o.payment_dt>=%(date_op)s
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.total_fee end))/100 as '本周',
        sum((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and o.payment_dt<=concat(DATE_SUB(%(date_ed)s, INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then o.total_fee end))/100 as '较前一周',
        sum((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and o.payment_dt<=concat(DATE_SUB(%(date_ed)s, INTERVAL 1 year),' 23:59:59') then o.total_fee end))/100 as '同期',
        sum((case when o.payment_dt>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.total_fee end))/100 as '本月'
        FROM hopsonone_order_center.orders o,
        hopsonone_order_center.orders_detail_plusvip p
        where o.order_type in(4,8)
        and  o.market_id=%(market_id)s
        and p.trade_no = o.trade_no
        and o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59')
        group by o.market_id
        """
        # 计时
        付费会员收益运行开始时间 = dt.datetime.now()
        付费会员收益 = pd.read_sql(付费会员收益sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        付费会员收益运行结束时间 = dt.datetime.now()
        付费会员收益运行sql用时 = (付费会员收益运行结束时间 - 付费会员收益运行开始时间).seconds
        print(f'付费会员收益运行SQL时间为{付费会员收益运行sql用时}秒')

        # 会员阶段数据SQL语句
        会员阶段数据sql = """
            select b.market_id,b.`日期`,b.`会员拉新`,c.`付费会员新增`,c.`付费会员收益`
        		from (SELECT  a.market_id,
            date_format(a.create_time,'%%m/%%d') as '日期',
            count(*) as '会员拉新'
            from hopsonone_members.members a
            where a.market_id=%(market_id)s
            and a.create_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
            and a.create_time<=concat(%(date_ed)s,' 23:59:59')
            GROUP BY a.market_id,date_format(a.create_time,'%%m/%%d')
            ORDER BY a.market_id,date_format(a.create_time,'%%m/%%d')) b,
        		(SELECT  o.market_id,
            date_format(o.payment_dt,'%%m/%%d') as '日期',
            count(o.buyer_id) as '付费会员新增',
            sum(o.total_fee)/100 as '付费会员收益'
            from hopsonone_order_center.orders_detail_plusvip a,
            hopsonone_order_center.orders o
            where a.trade_no = o.trade_no
            and o.market_id=%(market_id)s
            and o.state in (3, 4, 5)
            and o.order_type=4
            and o.payment_dt>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
            and o.payment_dt<=concat(%(date_ed)s,' 23:59:59')
            GROUP BY o.market_id,date_format(o.payment_dt,'%%m/%%d')
            ORDER BY o.market_id,date_format(o.payment_dt,'%%m/%%d')) c
        		where b.market_id=c.market_id
        		and b.`日期`=c.`日期`
            """
        # 计时
        会员阶段数据运行开始时间 = dt.datetime.now()
        会员阶段数据 = pd.read_sql(会员阶段数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        会员阶段数据运行结束时间 = dt.datetime.now()
        会员阶段数据运行sql用时 = (会员阶段数据运行结束时间 - 会员阶段数据运行开始时间).seconds
        print(f'会员阶段数据运行SQL时间为{会员阶段数据运行sql用时}秒')

        # 会员拉新渠道SQL语句
        会员拉新渠道sql = """
        SELECT a.market_id,
        s.dic_desc as '渠道',
        count(DISTINCT a.m_id) as '拉新人数'
        from hopsonone_members.members a,
        (SELECT  distinct s1.dic_value,
        (case when s1.dic_value in (0,8,15,10,9,21) then s1.dic_desc else '其他' end) as 'dic_desc'
        from hopsonone_cms.system_sys_dic s1
        where s1.type_name='REGISTER_TYPE') s
        where a.m_origin=s.dic_value
        and a.market_id=%(market_id)s
        and a.m_status = 1
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY a.market_id,s.dic_desc
        """
        # 计时
        会员拉新渠道运行开始时间 = dt.datetime.now()
        会员拉新渠道 = pd.read_sql(会员拉新渠道sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        会员拉新渠道运行结束时间 = dt.datetime.now()
        会员拉新渠道运行sql用时 = (会员拉新渠道运行结束时间 - 会员拉新渠道运行开始时间).seconds
        print(f'会员拉新渠道运行SQL时间为{会员拉新渠道运行sql用时}秒')

        # 付费会员开卡类型SQL语句
        付费会员开卡sql = """
        SELECT o.market_id,
        (case when a.plus_card_type=1 then '月卡'
        when a.plus_card_type!=1 then '年卡' end) as '类型',
        sum(case when a.opt_type=1 then 1 else 0 end ) as '开卡',
        sum(case when a.opt_type!=1 then 1 else 0 end ) as '续卡'
        from hopsonone_order_center.orders_detail_plusvip a,
        hopsonone_order_center.orders o
        where o.market_id=%(market_id)s
        and o.state in (3,4,5)
        and a.trade_no = o.trade_no
        and o.payment_dt>=%(date_op)s
        and o.payment_dt<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY o.market_id,a.plus_card_type
        order by o.market_id,a.plus_card_type
        """
        # 计时
        付费会员开卡运行开始时间 = dt.datetime.now()
        付费会员开卡 = pd.read_sql(付费会员开卡sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        付费会员开卡运行结束时间 = dt.datetime.now()
        付费会员开卡运行sql用时 = (付费会员开卡运行结束时间 - 付费会员开卡运行开始时间).seconds
        print(f'付费会员开卡运行SQL时间为{付费会员开卡运行sql用时}秒')

        # 线下销售SQL语句
        线下销售sql = """
        SELECT b.market_id,
        sum((case when a.trade_date>=%(date_op)s
        and a.trade_date<=%(date_ed)s then a.adjust_tax_sales_amount_month end))/100/10000 as '本周',
        sum((case when a.trade_date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.trade_date<=DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY) then a.adjust_tax_sales_amount_month end))/100/10000 as '较前一周',
        sum((case when a.trade_date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.trade_date<=DATE_SUB(%(date_ed)s,INTERVAL 1 year) then a.adjust_tax_sales_amount_month end))/100/10000 as '同期',
        sum((case when a.trade_date>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.trade_date<=%(date_ed)s then a.adjust_tax_sales_amount_month end))/100/10000 as '本月'
        FROM shop_side_operation_real_time.sales_report_details a,
        merchant_entity.entity_store b
        WHERE a.business_id = b.store_id
        and b.market_id=%(market_id)s
        and a.trade_date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.trade_date<=%(date_ed)s
        group by b.market_id
        """
        # 计时
        线下销售运行开始时间 = dt.datetime.now()
        线下销售 = pd.read_sql(线下销售sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        线下销售运行结束时间 = dt.datetime.now()
        线下销售运行sql用时 = (线下销售运行结束时间 - 线下销售运行开始时间).seconds
        print(f'线下销售运行SQL时间为{线下销售运行sql用时}秒')

        # 线下会员消费SQL语句
        线下会员消费sql = """
        SELECT p.market_id,
        IFNULL(sum((case when p.consume_time >=%(date_op)s
        and p.consume_time <=concat(%(date_ed)s,' 23:59:59') then p.consume_amount/100 end)),0)/10000 as '本周',
        IFNULL(sum((case when p.consume_time >=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and p.consume_time <=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then p.consume_amount/100 end)),0)/10000 as '较前一周',
        IFNULL(sum((case when p.consume_time >=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and p.consume_time <=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then p.consume_amount/100 end)),0)/10000 as '同期',
        IFNULL(sum((case when p.consume_time >=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and p.consume_time <=concat(%(date_ed)s,' 23:59:59') then p.consume_amount/100 end)),0)/10000 as '本月'
        FROM hopsonone_point_real_time.members_points_detail p
        where  p.market_id=%(market_id)s
        and p.state = 0
        and p.behavior in (8,20,24,26,27,28)
        and p.consume_time >=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and p.consume_time <=concat(%(date_ed)s,' 23:59:59')
        group by p.market_id
        """
        # 计时
        线下会员消费运行开始时间 = dt.datetime.now()
        线下会员消费 = pd.read_sql(线下会员消费sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        线下会员消费运行结束时间 = dt.datetime.now()
        线下会员消费运行sql用时 = (线下会员消费运行结束时间 - 线下会员消费运行开始时间).seconds
        print(f'线下会员消费运行SQL时间为{线下会员消费运行sql用时}秒')

        # 线下客流SQL语句
        线下客流sql = """
        select a.market_id,
        sum((case when a.date>=%(date_op)s
        and a.date<=%(date_ed)s then a.passenger_flow_count end))/10000 as '本周',
        sum((case when a.date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.date<=DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY) then a.passenger_flow_count end))/10000 as '较前一周',
        sum((case when a.date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.date<=DATE_SUB(%(date_ed)s,INTERVAL 1 year) then a.passenger_flow_count end))/10000 as '同期',
        sum((case when a.date>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.date<=%(date_ed)s then a.passenger_flow_count end))/10000 as '本月'
        from hopsonone_market_analysis.market_analysis_day a
        where a.market_id=%(market_id)s
        and a.date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.date<=%(date_ed)s
        GROUP BY a.market_id
        """
        # 计时
        线下客流运行开始时间 = dt.datetime.now()
        线下客流 = pd.read_sql(线下客流sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        线下客流运行结束时间 = dt.datetime.now()
        线下客流运行sql用时 = (线下客流运行结束时间 - 线下客流运行开始时间).seconds
        print(f'线下客流运行SQL时间为{线下客流运行sql用时}秒')

        # 线下车流SQL语句
        线下车流sql = """
        select a.market_id,
        sum((case when a.date>=%(date_op)s
        and a.date<=%(date_ed)s then ifnull(a.freight_flow_count,0) end)) as '本周',
        sum((case when a.date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.date<=DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY) then ifnull(a.freight_flow_count,0) end)) as '较前一周',
        sum((case when a.date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.date<=DATE_SUB(%(date_ed)s,INTERVAL 1 year) then ifnull(a.freight_flow_count,0) end)) as '同期',
        sum((case when a.date>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.date<=%(date_ed)s then ifnull(a.freight_flow_count,0) end)) as '本月'
        from hopsonone_market_analysis.market_analysis_day a
        where a.market_id=%(market_id)s
        and a.date>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.date<=%(date_ed)s
        GROUP BY a.market_id
        """
        # 计时
        线下车流运行开始时间 = dt.datetime.now()
        线下车流 = pd.read_sql(线下车流sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        线下车流运行结束时间 = dt.datetime.now()
        线下车流运行sql用时 = (线下车流运行结束时间 - 线下车流运行开始时间).seconds
        print(f'线下车流运行SQL时间为{线下车流运行sql用时}秒')

        # 线下消费阶段数据SQL语句
        线下消费阶段数据sql = """
        select a1.market_id,a1.`日期`,a1.`线下销售额（万元）`,a2.`线下会员消费（万元）`
        from
        (SELECT b.market_id,
        date_format(a.trade_date,'%%m/%%d') as '日期',
        sum(a.adjust_tax_sales_amount_month)/100/10000 as '线下销售额（万元）'
        FROM shop_side_operation_real_time.sales_report_details a,
        merchant_entity.entity_store b
        WHERE a.business_id=b.store_id
        and b.market_id=%(market_id)s
        and a.trade_date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.trade_date<=%(date_ed)s
        group by b.market_id,date_format(a.trade_date,'%%m/%%d')) a1,
        (SELECT p.market_id,
        date_format(p.consume_time,'%%m/%%d') as '日期',
        IFNULL(sum(p.consume_amount/100 ),0)/10000 as '线下会员消费（万元）'
        FROM hopsonone_point_real_time.members_points_detail p
        where  p.market_id=%(market_id)s
        and p.state = 0
        and p.behavior in (8,20,24,26,27,28)
        and p.consume_time >=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and p.consume_time <=concat(%(date_ed)s,' 23:59:59')
        group by p.market_id,date_format(p.consume_time,'%%m/%%d')) a2
        where a1.market_id=a2.market_id and a1.`日期`=a2.`日期`
        """
        # 计时
        线下消费阶段数据运行开始时间 = dt.datetime.now()
        线下消费阶段数据 = pd.read_sql(线下消费阶段数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                               index_col='market_id')
        线下消费阶段数据运行结束时间 = dt.datetime.now()
        线下消费阶段数据运行sql用时 = (线下消费阶段数据运行结束时间 - 线下消费阶段数据运行开始时间).seconds
        print(f'线下消费阶段数据运行SQL时间为{线下消费阶段数据运行sql用时}秒')

        # 线下流量阶段数据SQL语句
        线下流量阶段数据sql = """
        select a1.market_id,a1.`日期`,a1.`线下客流`,a2.`线下车流（辆）`
        from
        (select a.market_id,
        date_format(a.date,'%%m/%%d') as '日期',
        sum(a.passenger_flow_count) as '线下客流'
        from hopsonone_market_analysis.market_analysis_day a
        where a.market_id=%(market_id)s
        and a.date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.date<=%(date_ed)s
        GROUP BY a.market_id,date_format(a.date,'%%m/%%d')) a1,
        (select a.market_id,
        date_format(a.date,'%%m/%%d') as '日期1',
        sum(a.freight_flow_count) as '线下车流（辆）'
        from hopsonone_market_analysis.market_analysis_day a
        where a.market_id=%(market_id)s
        and a.date>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.date<=%(date_ed)s
        GROUP BY a.market_id,date_format(a.date,'%%m/%%d')) a2
        where a1.market_id=a2.market_id and a1.`日期`=a2.`日期1`
        """
        # 计时
        线下流量阶段数据运行开始时间 = dt.datetime.now()
        线下流量阶段数据 = pd.read_sql(线下流量阶段数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                               index_col='market_id')
        线下流量阶段数据运行结束时间 = dt.datetime.now()
        线下流量阶段数据运行sql用时 = (线下流量阶段数据运行结束时间 - 线下流量阶段数据运行开始时间).seconds
        print(f'线下流量阶段数据运行SQL时间为{线下流量阶段数据运行sql用时}秒')

        # 付费与非付费停车占比SQL语句
        付费与非付费停车占比sql = """
        select a1.market_id,a1.`付费车流量`,(a1.`总车流量`-a1.`付费车流量`) as'免费车流量'
        from
        (select a.market_id,
        count(*)as '总车流量',
        sum(case when a.pay_way=0 then 1 else 0 end) as '付费车流量'
        from hopsonone_park.park_order a
        where a.market_id=%(market_id)s
        and a.status in (3,4,5)
        and a.payment_time>=%(date_op)s
        and a.payment_time<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY a.market_id) a1
        """
        # 计时
        付费与非付费停车占比运行开始时间 = dt.datetime.now()
        付费与非付费停车占比 = pd.read_sql(付费与非付费停车占比sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                                 index_col='market_id')
        付费与非付费停车占比运行结束时间 = dt.datetime.now()
        付费与非付费停车占比运行sql用时 = (付费与非付费停车占比运行结束时间 - 付费与非付费停车占比运行开始时间).seconds
        print(f'付费与非付费停车占比运行SQL时间为{付费与非付费停车占比运行sql用时}秒')

        # 会员停车订单占比SQL语句
        会员停车订单占比sql = """
        select a1.market_id,(a1.`总停车数`-a2.`会员`) as '非会员',(a2.`会员`-a3.`付费会员`) as '普通会员',a3.`付费会员`
        from
        (select a.market_id,count(*) as '总停车数'
        from hopsonone_park.park_order a
        where a.market_id=%(market_id)s
        and a.payment_time>=%(date_op)s
        and a.payment_time<=concat(%(date_ed)s,' 23:59:59')
        and a.status in (3,4,5)
        GROUP BY a.market_id) a1,
        (select a.market_id,count(*) as '会员'
        from hopsonone_park.park_order a,
        hopsonone_members.members m
        where a.market_id=%(market_id)s
        and a.m_id=m.m_id
        and a.status in (3,4,5)
        and a.payment_time>=%(date_op)s
        and a.payment_time<=concat(%(date_ed)s,' 23:59:59')
        GROUP BY a.market_id) a2,
        (SELECT a.market_id,COUNT(*) AS '付费会员'
        FROM hopsonone_park.park_order a,
        hopsonone_personal.members_plusvip p
        WHERE a.market_id=%(market_id)s
        and a.status in (3,4,5)
        and a.m_id=p.m_id
        and a.payment_time>=%(date_op)s
        and a.payment_time<=concat(%(date_ed)s,' 23:59:59')
        and p.status=1
        and p.due_time >= a.payment_time
        and p.effective_time <= a.payment_time
        GROUP BY a.market_id) a3
        where a1.market_id=a2.market_id and a2.market_id=a3.market_id
        """
        # 计时
        会员停车订单占比运行开始时间 = dt.datetime.now()
        会员停车订单占比 = pd.read_sql(会员停车订单占比sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                               index_col='market_id')
        会员停车订单占比运行结束时间 = dt.datetime.now()
        会员停车订单占比运行sql用时 = (会员停车订单占比运行结束时间 - 会员停车订单占比运行开始时间).seconds
        print(f'会员停车订单占比运行SQL时间为{会员停车订单占比运行sql用时}秒')

        # 卡券收益SQL语句
        卡券收益sql = """
        select a.market_id,
        sum((case when a.receive_time>=%(date_op)s
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59') then a.amount end))/100 as '本周',
        sum((case when a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.receive_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.amount end))/100 as '较前一周',
        sum((case when a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.receive_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then a.amount end))/100 as '同期',
        sum((case when a.receive_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59') then a.amount end))/100 as '本月'
        from hopsonone_coupons_v2.coupons_sub_order a
        where a.market_id=%(market_id)s
        and a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
        and a.order_state in (3,4,5)
        and a.get_way=3
        group by a.market_id
        """
        # 计时
        卡券收益运行开始时间 = dt.datetime.now()
        卡券收益 = pd.read_sql(卡券收益sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        卡券收益运行结束时间 = dt.datetime.now()
        卡券收益运行sql用时 = (卡券收益运行结束时间 - 卡券收益运行开始时间).seconds
        print(f'卡券收益运行SQL时间为{卡券收益运行sql用时}秒')

        # 卡券领取SQL语句
        卡券领取sql = """
        select a.market_id,
        count((case when a.receive_time>=%(date_op)s
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59') then a.trade_no end)) as '本周',
        count((case when a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.receive_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.trade_no end)) as '较前一周',
        count((case when a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.receive_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then a.trade_no end)) as '同期',
        count((case when a.receive_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59') then a.trade_no end)) as '本月'
        from hopsonone_coupons_v2.coupons_sub_order a
        where  a.market_id=%(market_id)s
        and a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
        and a.order_state in (3,4,5)
        group by a.market_id
        """
        # 计时
        卡券领取运行开始时间 = dt.datetime.now()
        卡券领取 = pd.read_sql(卡券领取sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        卡券领取运行结束时间 = dt.datetime.now()
        卡券领取运行sql用时 = (卡券领取运行结束时间 - 卡券领取运行开始时间).seconds
        print(f'卡券领取运行SQL时间为{卡券领取运行sql用时}秒')

        # 卡券核销SQL语句
        卡券核销sql = """
        select a.market_id,
        count((case when a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.sub_trade_no end)) as '本周',
        count((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.sub_trade_no end)) as '较前一周',
        count((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then a.sub_trade_no end)) as '同期',
        count((case when a.create_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.sub_trade_no end)) as '本月'
        from hopsonone_coupons_v2.coupons_use_record a
        where  a.market_id=%(market_id)s
        and a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        group by a.market_id
        """
        # 计时
        卡券核销运行开始时间 = dt.datetime.now()
        卡券核销 = pd.read_sql(卡券核销sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        卡券核销运行结束时间 = dt.datetime.now()
        卡券核销运行sql用时 = (卡券核销运行结束时间 - 卡券核销运行开始时间).seconds
        print(f'卡券核销运行SQL时间为{卡券核销运行sql用时}秒')

        # 卡券阶段数据SQL语句
        卡券阶段数据sql = """
            select b.market_id,b.`日期`,b.`卡券领取`,b.`卡券收益（元）`,c.`卡券核销`
        		from (select a.market_id,
            date_format(a.receive_time,'%%m/%%d') as '日期',
            count(a.trade_no) as '卡券领取',
        		sum(case when a.get_way=3 then a.amount else 0 end)/100 as '卡券收益（元）'
            from hopsonone_coupons_v2.coupons_sub_order a
            where  a.market_id =%(market_id)s
            and a.receive_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
            and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
            and a.order_state in (3,4,5)
            group by a.market_id,date_format(a.receive_time,'%%m/%%d')
            ORDER BY a.market_id,date_format(a.receive_time,'%%m/%%d')) b,
        		(select a.market_id,
            date_format(a.create_time,'%%m/%%d') as '日期',
            count(a.sub_trade_no) as '卡券核销'
            from hopsonone_coupons_v2.coupons_use_record a
            where a.market_id=%(market_id)s
            and a.create_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
            and a.create_time<=concat(%(date_ed)s,' 23:59:59')
            group by a.market_id,date_format(a.create_time,'%%m/%%d')
            ORDER BY a.market_id,date_format(a.create_time,'%%m/%%d')) c
        		where b.market_id=c.market_id
        		and b.`日期`=c.`日期`
            """
        # 计时
        卡券阶段数据运行开始时间 = dt.datetime.now()
        卡券阶段数据 = pd.read_sql(卡券阶段数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        卡券阶段数据运行结束时间 = dt.datetime.now()
        卡券阶段数据运行sql用时 = (卡券阶段数据运行结束时间 - 卡券阶段数据运行开始时间).seconds
        print(f'卡券阶段数据运行SQL时间为{卡券阶段数据运行sql用时}秒')

        # 卡券领取渠道SQL语句
        卡券领取渠道sql = """
        select a.market_id,
        (case when a.get_channel_type=0 then '自动推送'
        when a.get_channel_type=1 then '卡券中心'
        when a.get_channel_type=2 then '超级会员'
        when a.get_channel_type=3 then '租户端口领取'
        when a.get_channel_type=4 then '游戏抽奖获取'
        when a.get_channel_type=5 then '推广码'
        when a.get_channel_type=6 then '导购屏'
        when a.get_channel_type=7 then '线下活动'
        when a.get_channel_type=9 then '推广码奖励'
        when a.get_channel_type=10 then '限时秒杀获取'
        when a.get_channel_type=11 then '粮票附赠'
        else '其他' end ) as '领取方式',
        count(a.trade_no) as '计数'
        from hopsonone_coupons_v2.coupons_sub_order a
        where a.market_id=%(market_id)s
        and a.order_state in (3,4,5)
        and a.receive_time>=%(date_op)s
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
        and a.order_state=3
        group by a.market_id,a.get_channel_type
        """
        # 计时
        卡券领取渠道运行开始时间 = dt.datetime.now()
        卡券领取渠道 = pd.read_sql(卡券领取渠道sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        卡券领取渠道运行结束时间 = dt.datetime.now()
        卡券领取渠道运行sql用时 = (卡券领取渠道运行结束时间 - 卡券领取渠道运行开始时间).seconds
        print(f'卡券领取渠道运行SQL时间为{卡券领取渠道运行sql用时}秒')

        # 卡券核销分类SQL语句
        卡券核销分类sql = """
        select a.market_id,
        (case when s.get_way=1 then '免费领取'
        when s.get_way=2 then '积分兑换'
        when s.get_way=3 then '现金支付' end) as '卡券核销分类',
        count(a.sub_trade_no) as '计数'
        from hopsonone_coupons_v2.coupons_use_record a,
        hopsonone_coupons_v2.coupons_sub_order s
        where a.market_id=%(market_id)s
        and a.coupons_number = s.coupons_number
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        group by a.market_id,s.get_way
        """
        # 计时
        卡券核销分类运行开始时间 = dt.datetime.now()
        卡券核销分类 = pd.read_sql(卡券核销分类sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        卡券核销分类运行结束时间 = dt.datetime.now()
        卡券核销分类运行sql用时 = (卡券核销分类运行结束时间 - 卡券核销分类运行开始时间).seconds
        print(f'卡券核销分类运行SQL时间为{卡券核销分类运行sql用时}秒')

        # 卡券排名数据SQL语句
        卡券排名数据sql = """
        select a1.market_id,a1.`卡券名称`,a1.`卡券收益（元）`,a1.`卡券领取`,a2.`卡券核销`
        from
        (select a.market_id,
        a.coupons_id,
        a.name as '卡券名称',
        sum(a.amount)/100 as '卡券收益（元）',
        count(a.trade_no) as '卡券领取'
        from hopsonone_coupons_v2.coupons_sub_order a
        where  a.market_id=%(market_id)s
        and a.order_state in (3,4,5)
        and a.get_way=3
        and a.receive_time>=%(date_op)s
        and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
        group by a.market_id,a.name) a1,
        (select b.market_id,b.coupons_id,b.coupons_name,count(b.sub_trade_no) as '卡券核销'
        from hopsonone_coupons_v2.coupons_use_record b
        where  b.market_id=%(market_id)s
        and b.create_time>=%(date_op)s
        and b.create_time<=concat(%(date_ed)s,' 23:59:59')
        group by b.market_id,b.coupons_name) a2
        where a1.market_id=a2.market_id and a1.coupons_id=a2.coupons_id
        order by a1.`卡券收益（元）` desc,a1.`卡券领取` desc,a2.`卡券核销`desc
        limit 10
        """
        # 计时
        卡券排名数据运行开始时间 = dt.datetime.now()
        卡券排名数据 = pd.read_sql(卡券排名数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        卡券排名数据运行结束时间 = dt.datetime.now()
        卡券排名数据运行sql用时 = (卡券排名数据运行结束时间 - 卡券排名数据运行开始时间).seconds
        print(f'卡券排名数据运行SQL时间为{卡券排名数据运行sql用时}秒')

        # 产生积分SQL语句
        产生积分sql = """
        select a.market_id,
        sum((case when a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.actual_point end))/10000 as '本周',
        sum((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.actual_point end))/10000 as '较前一周',
        sum((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then a.actual_point end))/10000 as '同期',
        sum((case when a.create_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.actual_point end))/10000 as '本月'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=0
        group by a.market_id
        """
        # 计时
        产生积分运行开始时间 = dt.datetime.now()
        产生积分 = pd.read_sql(产生积分sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        产生积分运行结束时间 = dt.datetime.now()
        产生积分运行sql用时 = (产生积分运行结束时间 - 产生积分运行开始时间).seconds
        print(f'产生积分运行SQL时间为{产生积分运行sql用时}秒')

        # 消耗积分SQL语句
        消耗积分sql = """
        select a.market_id,
        sum((case when a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.actual_point end))/10000 as '本周',
        sum((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY),' 23:59:59') then a.actual_point end))/10000 as '较前一周',
        sum((case when a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(DATE_SUB(%(date_ed)s,INTERVAL 1 year),' 23:59:59') then a.actual_point end))/10000 as '同期',
        sum((case when a.create_time>=concat(date_format(%(date_op)s,'%%Y-%%m'),'-01')
        and a.create_time<=concat(%(date_ed)s,' 23:59:59') then a.actual_point end))/10000 as '本月'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=DATE_SUB(%(date_op)s,INTERVAL 1 year)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=1
        group by a.market_id
        """
        # 计时
        消耗积分运行开始时间 = dt.datetime.now()
        消耗积分 = pd.read_sql(消耗积分sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                           index_col='market_id')
        消耗积分运行结束时间 = dt.datetime.now()
        消耗积分运行sql用时 = (消耗积分运行结束时间 - 消耗积分运行开始时间).seconds
        print(f'消耗积分运行SQL时间为{消耗积分运行sql用时}秒')

        # 积分阶段数据SQL语句
        积分阶段数据sql = """
        select a1.market_id,a1.`日期`,a1.`产生积分（万）`,a2.`消耗积分（万）`
        from
        (select a.market_id,
        date_format(a.create_time,'%%m/%%d') as '日期',
        sum(a.actual_point)/10000 as '产生积分（万）'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=date_sub(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=0
        group by a.market_id,date_format(a.create_time,'%%m/%%d')) a1,
        (select a.market_id,
        date_format(a.create_time,'%%m/%%d') as '日期',
        sum(a.actual_point)/10000 as '消耗积分（万）'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=date_sub(%(date_op)s,INTERVAL datediff(%(date_ed)s,%(date_op)s)+1 DAY)
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=1
        group by a.market_id,date_format(a.create_time,'%%m/%%d')) a2
        where a1.market_id=a2.market_id and a1.`日期`=a2.`日期`
        order by a1.market_id,a1.`日期`
        """
        # 计时
        积分阶段数据运行开始时间 = dt.datetime.now()
        积分阶段数据 = pd.read_sql(积分阶段数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        积分阶段数据运行结束时间 = dt.datetime.now()
        积分阶段数据运行sql用时 = (积分阶段数据运行结束时间 - 积分阶段数据运行开始时间).seconds
        print(f'积分阶段数据运行SQL时间为{积分阶段数据运行sql用时}秒')

        # 会员产生积分占比SQL语句
        会员产生积分占比sql = """
        select a1.market_id,(a1.`产生积分总数`-a2.`付费会员积分`) as '普通会员积分',a2.`付费会员积分`
        from
        (select a.market_id,
        sum(a.point) as '产生积分总数'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=0
        group by a.market_id) a1,
        (select a.market_id,
        sum(a.actual_point) as '付费会员积分'
        from hopsonone_point_real_time.members_points_detail a,
        hopsonone_personal.members_plusvip p
        where a.market_id=%(market_id)s
        and a.m_id = p.m_id
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=0
        and p.status= 1
        and p.due_time >= a.create_time
        and p.effective_time <= a.create_time
        group by a.market_id) a2
        where a1.market_id=a2.market_id
        group by a1.market_id

        """
        # 计时
        会员产生积分占比运行开始时间 = dt.datetime.now()
        会员产生积分占比 = pd.read_sql(会员产生积分占比sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                               index_col='market_id')
        会员产生积分占比运行结束时间 = dt.datetime.now()
        会员产生积分占比运行sql用时 = (会员产生积分占比运行结束时间 - 会员产生积分占比运行开始时间).seconds
        print(f'会员产生积分占比运行SQL时间为{会员产生积分占比运行sql用时}秒')

        # 会员产生消耗占比SQL语句
        会员产生消耗占比sql = """
        select a1.market_id,(a1.`消耗积分总数`-a2.`付费会员消耗积分`) as '普通会员消耗积分',a2.`付费会员消耗积分`
        from
        (select a.market_id,
        sum(a.point) as '消耗积分总数'
        from hopsonone_point_real_time.members_points_detail a
        where a.market_id=%(market_id)s
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=1
        group by a.market_id) a1,
        (select a.market_id,
        sum(a.point) as '付费会员消耗积分'
        from hopsonone_point_real_time.members_points_detail a,
        hopsonone_personal.members_plusvip p
        where  a.market_id=%(market_id)s
        and a.m_id = p.m_id
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=1
        and p.status= 1
        and p.due_time >= a.create_time
        and p.effective_time <= a.create_time
        group by a.market_id) a2
        where a1.market_id=a2.market_id
        group by a1.market_id
        """
        # 计时
        会员产生消耗占比运行开始时间 = dt.datetime.now()
        会员产生消耗占比 = pd.read_sql(会员产生消耗占比sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                               index_col='market_id')
        会员产生消耗占比运行结束时间 = dt.datetime.now()
        会员产生消耗占比运行sql用时 = (会员产生消耗占比运行结束时间 - 会员产生消耗占比运行开始时间).seconds
        print(f'会员产生消耗占比运行SQL时间为{会员产生消耗占比运行sql用时}秒')

        # 会员积分排名SQL语句
        会员积分排名sql = """
        select a.market_id,
        b.m_mobile as '会员手机号',
        sum(a.point) as '产生积分'
        from hopsonone_point_real_time.members_points_detail a,
        hopsonone_members.members b
        where a.m_id=b.m_id
        and a.market_id=%(market_id)s
        and a.create_time>=%(date_op)s
        and a.create_time<=concat(%(date_ed)s,' 23:59:59')
        and a.state=0
        group by a.market_id,b.m_mobile
        order by a.market_id,sum(a.point) desc
        limit 10
        """
        # 计时
        会员积分排名运行开始时间 = dt.datetime.now()
        会员积分排名 = pd.read_sql(会员积分排名sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期, 'market_id': 项目ID},
                             index_col='market_id')
        会员积分排名运行结束时间 = dt.datetime.now()
        会员积分排名运行sql用时 = (会员积分排名运行结束时间 - 会员积分排名运行开始时间).seconds
        print(f'会员积分排名运行SQL时间为{会员积分排名运行sql用时}秒')

        # 取读模板
        运营周报 = openpyxl.load_workbook('week_stats.xlsx')

        周报 = 运营周报['周报']
        if 会员拉新.shape[0] > 0:
            周报['R4'].value = 会员拉新.loc[项目ID]['本周']
            周报['S4'].value = 会员拉新.loc[项目ID]['本月']
            周报['T4'].value = 会员拉新.loc[项目ID]['较前一周']
            周报['U4'].value = 会员拉新.loc[项目ID]['同期']
        if 付费会员新增.shape[0] > 0:
            周报['R7'].value = 付费会员新增.loc[项目ID]['本周']
            周报['S7'].value = 付费会员新增.loc[项目ID]['本月']
            周报['T7'].value = 付费会员新增.loc[项目ID]['较前一周']
            周报['U7'].value = 付费会员新增.loc[项目ID]['同期']
        if 付费会员收益.shape[0] > 0:
            周报['R10'].value = 付费会员收益.loc[项目ID]['本周']
            周报['S10'].value = 付费会员收益.loc[项目ID]['本月']
            周报['T10'].value = 付费会员收益.loc[项目ID]['较前一周']
            周报['U10'].value = 付费会员收益.loc[项目ID]['同期']
        if 线下销售.shape[0] > 0:
            周报['R52'].value = 线下销售.loc[项目ID]['本周']
            周报['S52'].value = 线下销售.loc[项目ID]['本月']
            周报['T52'].value = 线下销售.loc[项目ID]['较前一周']
            周报['U52'].value = 线下销售.loc[项目ID]['同期']
        if 线下会员消费.shape[0] > 0:
            周报['R55'].value = 线下会员消费.loc[项目ID]['本周']
            周报['S55'].value = 线下会员消费.loc[项目ID]['本月']
            周报['T55'].value = 线下会员消费.loc[项目ID]['较前一周']
            周报['U55'].value = 线下会员消费.loc[项目ID]['同期']
        if 线下客流.shape[0] > 0:
            周报['R58'].value = 线下客流.loc[项目ID]['本周']
            周报['S58'].value = 线下客流.loc[项目ID]['本月']
            周报['T58'].value = 线下客流.loc[项目ID]['较前一周']
            周报['U58'].value = 线下客流.loc[项目ID]['同期']
        if 线下车流.shape[0] > 0:
            周报['R61'].value = 线下车流.loc[项目ID]['本周']
            周报['S61'].value = 线下车流.loc[项目ID]['本月']
            周报['T61'].value = 线下车流.loc[项目ID]['较前一周']
            周报['U61'].value = 线下车流.loc[项目ID]['同期']
        if 付费与非付费停车占比.shape[0] > 0:
            周报['R101'].value = 付费与非付费停车占比.loc[项目ID]['付费车流量']
            周报['S101'].value = 付费与非付费停车占比.loc[项目ID]['免费车流量']
        if 会员停车订单占比.shape[0] > 0:
            周报['R104'].value = 会员停车订单占比.loc[项目ID]['非会员']
            周报['S104'].value = 会员停车订单占比.loc[项目ID]['普通会员']
            周报['T104'].value = 会员停车订单占比.loc[项目ID]['付费会员']
        if 卡券收益.shape[0] > 0:
            周报['R112'].value = 卡券收益.loc[项目ID]['本周']
            周报['S112'].value = 卡券收益.loc[项目ID]['本月']
            周报['T112'].value = 卡券收益.loc[项目ID]['较前一周']
            周报['U112'].value = 卡券收益.loc[项目ID]['同期']
        if 卡券领取.shape[0] > 0:
            周报['R115'].value = 卡券领取.loc[项目ID]['本周']
            周报['S115'].value = 卡券领取.loc[项目ID]['本月']
            周报['T115'].value = 卡券领取.loc[项目ID]['较前一周']
            周报['U115'].value = 卡券领取.loc[项目ID]['同期']
        if 卡券核销.shape[0] > 0:
            周报['R118'].value = 卡券核销.loc[项目ID]['本周']
            周报['S118'].value = 卡券核销.loc[项目ID]['本月']
            周报['T118'].value = 卡券核销.loc[项目ID]['较前一周']
            周报['U118'].value = 卡券核销.loc[项目ID]['同期']
        if 产生积分.shape[0] > 0:
            周报['R166'].value = 产生积分.loc[项目ID]['本周']
            周报['S166'].value = 产生积分.loc[项目ID]['本月']
            周报['T166'].value = 产生积分.loc[项目ID]['较前一周']
            周报['U166'].value = 产生积分.loc[项目ID]['同期']
        if 消耗积分.shape[0] > 0:
            周报['R169'].value = 消耗积分.loc[项目ID]['本周']
            周报['S169'].value = 消耗积分.loc[项目ID]['本月']
            周报['T169'].value = 消耗积分.loc[项目ID]['较前一周']
            周报['U169'].value = 消耗积分.loc[项目ID]['同期']
        if 会员产生积分占比.shape[0] > 0:
            周报['R193'].value = 会员产生积分占比.loc[项目ID]['普通会员积分']
            周报['S193'].value = 会员产生积分占比.loc[项目ID]['付费会员积分']
        if 会员产生消耗占比.shape[0] > 0:
            周报['R196'].value = 会员产生消耗占比.loc[项目ID]['普通会员消耗积分']
            周报['S196'].value = 会员产生消耗占比.loc[项目ID]['付费会员消耗积分']
        for a in range(付费会员开卡.shape[0]):
            for b in range(付费会员开卡.shape[1]):
                周报.cell(44 + a, 18 + b).value = 付费会员开卡.iloc[a][b]
        if 会员阶段数据.shape[0] > 0:
            for i in range(len(会员阶段数据['日期'])):
                周报.cell(15 + i, 18).value = 会员阶段数据.iloc[i][0]
                周报.cell(15 + i, 19).value = 会员阶段数据.iloc[i][1]
                周报.cell(15 + i, 20).value = 会员阶段数据.iloc[i][2]
                周报.cell(15 + i, 21).value = 会员阶段数据.iloc[i][3]
        if 线下消费阶段数据.shape[0] > 0:
            for d in range(len(线下消费阶段数据['日期'])):
                周报.cell(67 + d, 18).value = 线下消费阶段数据.iloc[d][0]
                周报.cell(67 + d, 19).value = 线下消费阶段数据.iloc[d][1]
                周报.cell(67 + d, 20).value = 线下消费阶段数据.iloc[d][2]
        if 线下流量阶段数据.shape[0] > 0:
            for e in range(len(线下流量阶段数据['日期'])):
                周报.cell(84 + e, 18).value = 线下流量阶段数据.iloc[e][0]
                周报.cell(84 + e, 19).value = 线下流量阶段数据.iloc[e][1]
                周报.cell(84 + e, 20).value = 线下流量阶段数据.iloc[e][2]
        if 卡券阶段数据.shape[0] > 0:
            for f in range(len(卡券阶段数据['日期'])):
                周报.cell(122 + f, 18).value = 卡券阶段数据.iloc[f][0]
                周报.cell(122 + f, 19).value = 卡券阶段数据.iloc[f][1]
                周报.cell(122 + f, 20).value = 卡券阶段数据.iloc[f][3]
                周报.cell(122 + f, 21).value = 卡券阶段数据.iloc[f][2]
        if 积分阶段数据.shape[0] > 0:
            for g in range(len(积分阶段数据['日期'])):
                周报.cell(176 + g, 18).value = 积分阶段数据.iloc[g][0]
                周报.cell(176 + g, 19).value = 积分阶段数据.iloc[g][1]
                周报.cell(176 + g, 20).value = 积分阶段数据.iloc[g][2]
        if 会员拉新渠道.shape[0] > 0:
            for j in range(len(会员拉新渠道['渠道'])):
                周报.cell(33 + j, 18).value = 会员拉新渠道.iloc[j][0]
                周报.cell(33 + j, 19).value = 会员拉新渠道.iloc[j][1]
        if 卡券领取渠道.shape[0] > 0:
            for n in range(len(卡券领取渠道['领取方式'])):
                周报.cell(138 + n, 18).value = 卡券领取渠道.iloc[n][0]
                周报.cell(138 + n, 19).value = 卡券领取渠道.iloc[n][1]
        if 卡券核销分类.shape[0] > 0:
            for o in range(len(卡券核销分类['卡券核销分类'])):
                周报.cell(138 + o, 21).value = 卡券核销分类.iloc[o][0]
                周报.cell(138 + o, 22).value = 卡券核销分类.iloc[o][1]
        if 卡券排名数据.shape[0] > 0:
            for t in range(len(卡券排名数据['卡券名称'])):
                周报.cell(150 + t, 18).value = 卡券排名数据.iloc[t][0]
                周报.cell(150 + t, 19).value = 卡券排名数据.iloc[t][1]
                周报.cell(150 + t, 20).value = 卡券排名数据.iloc[t][3]
                周报.cell(150 + t, 21).value = 卡券排名数据.iloc[t][2]
        if 会员积分排名.shape[0] > 0:
            for s in range(len(会员积分排名['会员手机号'])):
                周报.cell(204 + s, 18).value = 会员积分排名.iloc[s][0]
                周报.cell(204 + s, 19).value = 会员积分排名.iloc[s][1]

        周报['B1'].value = 项目名称 + '周报(' + 周报日期 + ')'

        # 重命名
        新路径 = os.path.join(output_dir+项目名称 + '周报' + 周报日期, 项目名称 + '周报' + 周报日期 + '.xlsx')
        运营周报.save(新路径)
        print(f'{项目名称}{周报日期}周报已生成')

if __name__ == '__main__':

    # add by ma.fei
    output_dir = './out/运营周报/'
    os.system('rm -rf ./out/运营周报/*')

    start_rq, end_rq = get_monday_to_sunday(datetime.now().strftime('%Y-%m-%d'), -1)
    week_rq = start_rq[5:7] + start_rq[8:] + '~' + end_rq[5:7] + end_rq[8:]

    print('合生通运营（周报）:')
    print('-----------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('周报日期：', week_rq)
    print('-----------------------------')

    # 生成各项目销售积分&运营积分(周报)
    stats_sales_points(start_rq,end_rq,week_rq)
    print('stats_sales_points exec complete!')

    # 各项目运营情况（周报）
    stats_week_report(start_rq,end_rq,week_rq)
    print('stats_week_report exec complete!')

    # 文件打包
    path_original = 'out/运营周报'
    path_zip='./out/运营周报/运营周报-{}.zip'.format(week_rq)
    zip_name ='运营周报-{}.zip'.format(week_rq)

    file_to_zip(path_original,path_zip)
    print('文件打包完成!')

    # 发送邮件及附件
    sender='190634@lifeat.cn,850646@cre-hopson.com,820618@cre-hopson.com,546564@hopson.com.cn'
    Cc = '190343@lifeat.cn,820987@cre-hopson.com'
    title='合生通运营周报-{}'.format(week_rq)
    content='''各位领导：
    {}
    详见附件,请查收。
    '''.format(title)
    # 附件列表
    os.chdir('./out/运营周报')
    file_list = [zip_name]
    ret = SendMail(sender, Cc,title, content).send(file_list)
    print(ret,type(ret))
    print('邮件邮件完成')