#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os
import re
import pymysql
import pymongo
import pymysql as py
import pandas as pd
import datetime as dt
import openpyxl
import zipfile
import pathlib
import smtplib
from openpyxl.styles import Side,Border,PatternFill,Font
from datetime import datetime,timedelta
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

class SendMail(object):
    '''
    https://cloud.tencent.com/developer/article/1598257
    '''
    def __init__(self,sender,Cc,title,content):
        self.sender = sender.split(',')  #发送地址
        self.Cc = Cc.split(',')  # 抄送地址
        self.title = title  # 标题
        self.content = content  # 发送内容
        self.sys_sender = '190343@lifeat.cn'  # 系统账户
        self.sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码

    def send(self,file_list):
        """
        发送邮件
        :param file_list: 附件文件列表
        :return: bool
        """
        try:
            # 创建一个带附件的实例
            msg = MIMEMultipart()
            # 发件人格式
            msg['From'] = formataddr(["技术服务部", self.sys_sender])
            # 收件人格式
            print('a=',[formataddr(["", u]) for u in self.sender])
            msg['To'] = sender

            # 抄送人
            msg['Cc'] = Cc

            # 邮件主题
            msg['Subject'] = self.title

            # 邮件正文内容
            msg.attach(MIMEText(self.content, 'plain', 'utf-8'))

            # 多个附件
            for file_name in file_list:
                print("file_name",file_name)
                # 构造附件
                xlsxpart = MIMEApplication(open(file_name, 'rb').read())
                # filename表示邮件中显示的附件名
                xlsxpart.add_header('Content-Disposition','attachment',filename = '%s'%file_name)
                msg.attach(xlsxpart)

            # SMTP服务器
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465,timeout=10)
            # 登录账户
            server.login(self.sys_sender, self.sys_pwd)
            # 发送邮件
            #server.sendmail(self.sys_sender, [self.sender, ], msg.as_string())
            server.sendmail(self.sys_sender,
                            [formataddr(["", u]) for u in self.sender]+
                            [formataddr(["", u]) for u in self.Cc],
                            msg.as_string())
            # 退出账户
            server.quit()
            return True
        except Exception as e:
            print(e)
            return False

def file_to_zip(path_original, path_zip):
    '''
     作用：压缩文件到指定压缩包里
     参数一：压缩文件的位置
     参数二：压缩后的压缩包
    '''
    # 提前读取，避免把压缩包自己加上去
    # 这里用list()做一个克隆提前执行下，不然会在后面循环时才执行这一引用，如果压缩包在这个路径下，会将它读取进来。
    f_list = list(pathlib.Path(path_original).glob("**/*"))
    z = zipfile.ZipFile(path_zip, 'w')
    #print('len(path_original)=',len(path_original))
    for f in f_list:
        #print('f=',f,str(f),str(f)[len(path_original):])
        z.write(f, str(f)[len(path_original):])
    z.close()

def get_monday_to_sunday(today, weekly=0):
    """
    :function: 获取指定日期的周一和周日的日期
    :param today: '2021-11-16'; 当前日期：today = datetime.now().strftime('%Y-%m-%d')
    :param weekly: 获取指定日期的上几周或者下几周，weekly=0当前周，weekly=-1上一周，weekly=1下一周
    :return: 返回指定日期的周一和周日日期
    :return_type: tuple
    """
    last = weekly * 7
    today = datetime.strptime(str(today), "%Y-%m-%d")
    monday = datetime.strftime(today - timedelta(today.weekday() - last), "%Y-%m-%d")
    monday_ = datetime.strptime(monday, "%Y-%m-%d")
    sunday = datetime.strftime(monday_ + timedelta(monday_.weekday() + 6), "%Y-%m-%d")
    return monday, sunday

def format_sql(v_sql):
    return v_sql.replace("\\","\\\\").replace("'","\\'")

def get_ds_mysql(ip,port,service ,user,password):
    conn = pymysql.connect(host=ip,
                           port=int(port),
                           user=user,
                           passwd=password,
                           db=service,
                           charset='utf8',
                           cursorclass = pymysql.cursors.DictCursor,
                           autocommit=False)
    return conn

def get_ds_mongo(mongodb_str):
    ip            = mongodb_str.split(':')[0]
    port          = mongodb_str.split(':')[1]
    service       = mongodb_str.split(':')[2]
    user          = mongodb_str.split(':')[3]
    password      = mongodb_str.split(':')[4]
    conn          = pymongo.MongoClient(host=ip, port=int(port))
    db            = conn[service]
    db.authenticate(user, password)
    return db

def query_pos_from_mongo():
    db_mongo = get_ds_mongo('dds-2ze39d52c51a85f42163-pub.mongodb.rds.aliyuncs.com:3717:entity_store:hopsonone_ro:lvOD4ljLBkREapZd')
    rs_mongo=db_mongo['entity_dldf'].aggregate([{'$match':{'correlationStatus':3}},{'$project':{'_id':0,'businessId':1,'rentName':1,'cashName':1}}])
    return [ tuple(s.values())  for s in rs_mongo]

def ins_pos_to_mysql():
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', '3306', 'hopsonone_do', 'hopsonone_do', '8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    cr_mysql.execute('truncate table pos_mongo')
    print('Table `pos_mongo` truncate!')
    ins_data = query_pos_from_mongo()
    ins_tpl = 'insert into pos_mongo(`businessId`,`rentName`,`cashName`) values (%s,%s,%s)'
    cr_mysql.executemany(ins_tpl, ins_data)
    print('Table `pos_mongo` write complete!')
    db_mysql.commit()

def exec_proc_pos_source(p_proc_name,p_rq_begin,p_rq_end):
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com','3306','hopsonone_do','hopsonone_do','8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    print('call {}...'.format(p_proc_name))
    cr_mysql.callproc(p_proc_name,args=(p_rq_begin,p_rq_end))
    print('call {}...complete'.format(p_proc_name))

def exec_proc_pos_hz(p_proc_name):
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com','3306','hopsonone_do','hopsonone_do','8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    print('call {}...'.format(p_proc_name))
    cr_mysql.callproc(p_proc_name,)
    print('call {}...complete'.format(p_proc_name))

def upd_pos_mongo():
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', '3306', 'hopsonone_do', 'hopsonone_do','8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    cr_mysql.execute('select * from pos_数据源')
    rs= cr_mysql.fetchall()
    rs_desc = cr_mysql.description
    for r in rs:
       if r['项目ID'] == 218 and r['铺位号'] is not None \
           and len(re.findall('B[1-2]-[0-9]{2,3}', r['铺位号'])) > 0 \
               and int(re.findall('B[1-2]-[0-9]{2,3}', r['铺位号'])[0].split('-')[1]) > 57:
            r['楼层'] = r['楼层'] + '街区'

    cr_mysql.execute('delete from `pos_数据源`')
    print('Table `pos_数据源`表数据清空!')
    ins_data = [tuple(s.values()) for s in rs]
    ins_header="insert into `pos_数据源`( `项目ID`,`商场名称`,`商家编码`,`商家名称`,`租金架构`,`收银方式（合同）`,`收银方式（实际）`,`业态`,`楼层`,`铺位号`,`商户是否已开通POS`,`是否是海信POS`,`商户业绩上报营业额`,`交易笔数`,`非现金收入(元)`,`非现金退款(元)`,`非现金净收入(元)`,`现金收入(元)`,`现金退款(元)`,`现金净收入(元)`,`POS总收入(元)`,`POS总净收入(元)`,`非现金收入(元)/商户业绩上报营业额(%)`,`非现金净收入/商户业绩上报营业额(%)`)  values "
    ins_sql = ''
    for i in range(len(ins_data)):
        t = ""
        for j in range(len(ins_data[i])):
            col_type = str(rs_desc[j][1])
            if ins_data[i][j] is None:
                t = t + "null,"
            elif col_type == '253':  # varchar,date
                t = t + "'" + format_sql(str(ins_data[i][j])) + "',"
            elif col_type in ('1', '3', '8', '246'):  # int,decimal
                t = t +  str(ins_data[i][j]) + ","
            elif col_type == '12':  # datetime
                t = t + "'" + str(ins_data[i][j]).split('.')[0] + "',"
            else:
                t = t + "'" + format_sql(str(ins_data[i][j])) + "',"
        ins_sql = ins_sql + '(' + t[0:-1] + '),'
    cr_mysql.execute(ins_header+ins_sql[0:-1])
    print('Table `pos_数据源` update complete!')
    db_mysql.commit()

def write_pos_week(p_rq_start,p_rq_end,p_week_rq):
    开始日期, 结束日期, 周报日期 = p_rq_start, p_rq_end, p_week_rq
    date1 = dt.datetime.strptime(开始日期, '%Y-%m-%d').date()
    date2 = dt.datetime.strptime(结束日期, '%Y-%m-%d').date()

    # 连接数据库，执行SQL语句
    连接对象 = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',
                      passwd='8Loxk2IJxaenJkE3', database='hopsonone_do')

    项目文件夹 = output_dir + 'POS周报-' + 周报日期
    print('项目文件夹=',项目文件夹)
    os.mkdir(项目文件夹)

    # pos未上报业绩商户POS流水明细sql语句
    pos未上报业绩商户POS流水明细sql = """ select * from hopsonone_do.pos_未上报业绩商户POS流水明细"""
    pos未上报业绩商户POS流水明细 = pd.read_sql(pos未上报业绩商户POS流水明细sql, con=连接对象)

    # 商户业绩上报数据SQL语句
    pos数据源sql = """ select * from hopsonone_do.pos_数据源"""
    pos数据源 = pd.read_sql(pos数据源sql, con=连接对象)

    # 项目汇总SQL语句
    项目汇总sql = """ select * from hopsonone_do.pos_汇总"""
    项目汇总 = pd.read_sql(项目汇总sql, con=连接对象)

    # 合生通项目组数据SQL语句
    合生通项目组数据sql = """
    select a.*,e.`非现金收入(万元)`,e.`非现金退款(万元)`,e.`非现金净收入（万元）`
    from (SELECT 
    case b.status
    when 1 then '启用'
    when 0 then '禁用'
    end as '状态',
    '合生通项目组' as '项目名称',
    b.business_id as '商户编号',
    b.business_name as '商户名称'
    FROM hopson_hft.business b,
    hopson_hft.business_terminal t
    where b.business_id = t.business_id
    and b.market_id=0
    group by b.business_id
    order by b.business_id) a
    left join
    (select c.`非现金收入(万元)`,c.`非现金退款(万元)`,
    c.`非现金收入(万元)`-c.`非现金退款(万元)` as '非现金净收入（万元）',
    c.`商户编号`
    from 
    (SELECT
    o.business_id as '商户编号',
    round(ifnull(sum(case when o.direction_status=1 then o.total_amount END)/100/10000,0),2) as '非现金收入(万元)',
    round(ifnull(sum(case when o.direction_status=2 then o.total_amount END)/100/10000,0),2) as '非现金退款(万元)'
    FROM hopson_hft_real_time.intel_order o,
    hopson_hft_real_time.intel_order_payment p
    where o.market_id=0
    and p.sub_code = 1 
    and p.trans_way != 3
    and o.trade_no = p.trade_no
    and o.create_time >=%(date_op)s
    and o.create_time <= concat(%(date_ed)s,' 23:59:59')
    GROUP BY o.business_id) c
    order by c.`商户编号`) e
    on a.`商户编号`=e.`商户编号` collate utf8mb4_unicode_ci
    """
    合生通项目组 = pd.read_sql(合生通项目组数据sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期})

    数据源拆分列 = list(pos数据源['商场名称'].drop_duplicates())
    for i in 数据源拆分列:
        pos新数据源 = pos数据源[pos数据源['商场名称'] == i]
        pos未上报新数据源 = pos未上报业绩商户POS流水明细[pos未上报业绩商户POS流水明细['商场名称'] == i]
        项目汇总新 = 项目汇总[项目汇总['商场名称'] == i]
        if (date2 - date1).days > 10:
            with pd.ExcelWriter(项目文件夹 + '/' + i + '-POS月报' + 周报日期 + '.xlsx') as 表:
                项目汇总新.to_excel(表, sheet_name='汇总', index=False)
                pos新数据源.to_excel(表, sheet_name='数据明细', index=False)
                pos未上报新数据源.to_excel(表, sheet_name='pos未上报业绩商户POS流水明细', index=False)
        else:
            with pd.ExcelWriter(项目文件夹 + '/' + i + '-POS周报' + 周报日期 + '.xlsx') as 表:
                项目汇总新.to_excel(表, sheet_name='汇总', index=False)
                pos新数据源.to_excel(表, sheet_name='数据明细', index=False)
                pos未上报新数据源.to_excel(表, sheet_name='pos未上报业绩商户POS流水明细', index=False)

        # 取读模板
        if (date2 - date1).days > 10:
            路径 = 项目文件夹 + '/' + i + '-POS月报' + 周报日期 + '.xlsx'
        else:
            路径 = 项目文件夹 + '/' + i + '-POS周报' + 周报日期 + '.xlsx'

        原始表 = openpyxl.load_workbook(路径)
        原始汇总表 = 原始表['汇总']
        行 = 原始汇总表.max_row
        线 = Side(style="thin", color="000000")
        框 = Border(left=线, right=线, top=线, bottom=线)
        for r in 原始汇总表['A1':'I' + str(行)]:
            for c in r:
                c.border = 框
        色 = PatternFill(patternType="solid", start_color='D9E1F2')
        粗 = Font(bold=True)
        for k in 原始汇总表['A1:I1']:
            for t in k:
                t.fill = 色
        for m in 原始汇总表['A' + str(行):'I' + str(行)]:
            for n in m:
                n.fill = 色
                n.font = 粗
        原始汇总表.column_dimensions["A"].width = 20
        原始汇总表.column_dimensions["B"].width = 22
        原始汇总表.column_dimensions["C"].width = 20
        原始汇总表.column_dimensions["D"].width = 28
        原始汇总表.column_dimensions["E"].width = 15
        原始汇总表.column_dimensions["F"].width = 19
        原始汇总表.column_dimensions["G"].width = 37
        原始汇总表.column_dimensions["H"].width = 22
        原始汇总表.column_dimensions["I"].width = 37
        原始表.save(filename=路径)

    合生通项目组.to_excel(项目文件夹 + '/' + '合生通项目组-POS周报' + 周报日期 + '.xlsx', index=False)

if __name__ == '__main__':
    output_dir = './out/POS周报/'
    os.system('rm -rf ./out/POS周报/*')
    start_rq, end_rq = get_monday_to_sunday(datetime.now().strftime('%Y-%m-%d'), -1)
    week_rq = start_rq[5:7] + start_rq[8:] + '~' + end_rq[5:7] + end_rq[8:]

    print('合生通项目组-POS周报:')
    print('-----------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('周报日期：', week_rq)
    print('-----------------------------')

    # 1. 读mongo表写入mysql
    ins_pos_to_mysql()

    # 2.执行存储过程proc_week_pos_db_source预处理
    exec_proc_pos_source('proc_week_pos_source',start_rq,end_rq)

    # 3.更新pos_mongo表街区
    upd_pos_mongo()

    # 4.执行存储过程proc_week_pos_hz预处理
    exec_proc_pos_hz('proc_week_pos_hz')

    # 5.生成pos周报
    write_pos_week(start_rq,end_rq,week_rq)

    # 文件打包
    path_original = 'out/POS周报'
    path_zip = './out/POS周报/POS周报-{}.zip'.format(week_rq)
    zip_name = 'POS周报-{}.zip'.format(week_rq)

    file_to_zip(path_original, path_zip)
    print('文件打包完成!')

    # 发送邮件及附件
    sender = '190634@lifeat.cn,548400@lifeat.cn'
    Cc='190343@lifeat.cn,820987@cre-hopson.com'
    title = '合生通POS周报-{}'.format(week_rq)
    content = '''各位领导：
       {}
       详见附件,请查收。
       '''.format(title)
    # 附件列表
    os.chdir('./out/POS周报')
    file_list = [zip_name]
    ret = SendMail(sender,Cc,title, content).send(file_list)
    print(ret, type(ret))
    print('邮件邮件完成')