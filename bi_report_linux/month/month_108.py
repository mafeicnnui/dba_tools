#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/1/3 9:02
# @Author : ma.fei
# @File : month_108.py.py
# @Software: PyCharm

import pymysql
import pandas as pd
import datetime as dt
import os
import openpyxl
import zipfile
import pathlib
import smtplib
from datetime import datetime,timedelta
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

class SendMail(object):
    '''
    https://cloud.tencent.com/developer/article/1598257
    '''
    def __init__(self,sender,Cc,title,content):
        self.sender = sender.split(',')  #发送地址
        self.Cc = Cc.split(',')  # 抄送地址
        self.title = title  # 标题
        self.content = content  # 发送内容
        self.sys_sender = '190343@lifeat.cn'  # 系统账户
        self.sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码

    def send(self,file_list):
        """
        发送邮件
        :param file_list: 附件文件列表
        :return: bool
        """
        try:
            # 创建一个带附件的实例
            msg = MIMEMultipart()
            # 发件人格式
            msg['From'] = formataddr(["技术服务部", self.sys_sender])
            # 收件人格式
            print('a=',[formataddr(["", u]) for u in self.sender])
            msg['To'] = sender

            # 抄送人
            msg['Cc'] = Cc

            # 邮件主题
            msg['Subject'] = self.title

            # 邮件正文内容
            msg.attach(MIMEText(self.content, 'plain', 'utf-8'))

            # 多个附件
            for file_name in file_list:
                print("file_name",file_name)
                # 构造附件
                xlsxpart = MIMEApplication(open(file_name, 'rb').read())
                # filename表示邮件中显示的附件名
                xlsxpart.add_header('Content-Disposition','attachment',filename = '%s'%file_name)
                msg.attach(xlsxpart)

            # SMTP服务器
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465,timeout=10)
            # 登录账户
            server.login(self.sys_sender, self.sys_pwd)
            # 发送邮件
            #server.sendmail(self.sys_sender, [self.sender, ], msg.as_string())
            server.sendmail(self.sys_sender,
                            [formataddr(["", u]) for u in self.sender]+
                            [formataddr(["", u]) for u in self.Cc],
                            msg.as_string())
            # 退出账户
            server.quit()
            return True
        except Exception as e:
            print(e)
            return False

def file_to_zip(path_original, path_zip):
    '''
     作用：压缩文件到指定压缩包里
     参数一：压缩文件的位置
     参数二：压缩后的压缩包
    '''
    # 提前读取，避免把压缩包自己加上去
    # 这里用list()做一个克隆提前执行下，不然会在后面循环时才执行这一引用，如果压缩包在这个路径下，会将它读取进来。
    f_list = list(pathlib.Path(path_original).glob("**/*"))
    z = zipfile.ZipFile(path_zip, 'w')
    #print('len(path_original)=',len(path_original))
    for f in f_list:
        #print('f=',f,str(f),str(f)[len(path_original):])
        z.write(f, str(f)[len(path_original):])
    z.close()

def get_last_month_first_last_day():
    now = datetime.now()
    this_month_start = datetime(now.year, now.month, 1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    v_last_month_start = datetime.strftime(last_month_start, "%Y-%m-%d")
    v_last_month_end = datetime.strftime(last_month_end, "%Y-%m-%d")
    return v_last_month_start, v_last_month_end

def format_sql(v_sql):
    return v_sql.replace("\\","\\\\").replace("'","\\'")

def get_ds_mysql(ip,port,service ,user,password):
    conn = pymysql.connect(host=ip,
                           port=int(port),
                           user=user,
                           passwd=password,
                           db=service,
                           charset='utf8',
                           cursorclass = pymysql.cursors.DictCursor,
                           autocommit=True)
    return conn

def exec_proc(p_proc_name,p_rq_begin,p_rq_end):
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com','3306','hopsonone_do','hopsonone_do','8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    print('call {}...'.format(p_proc_name))
    cr_mysql.callproc(p_proc_name,args=(p_rq_begin,p_rq_end))
    print('call {}...complete'.format(p_proc_name))

def write_month_108(p_rq_start,p_rq_end,p_week_rq):
    开始日期, 结束日期, 周报日期 = p_rq_start, p_rq_end, p_week_rq
    # 取读模板
    项目月度总结 = openpyxl.load_workbook('month_108.xlsx')
    # 连接数据库，执行SQL语句
    连接对象 = pymysql.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',
                        passwd='8Loxk2IJxaenJkE3', database='hopsonone_do')
    # 会员注册量SQL语句
    累计会员量sql = """
    SELECT ma.market_name as '商场名称',
    sum(case when a.sex=1 then 1 end) as '男',
    sum(case when a.sex=2 then 1 end) as '女',
    sum(case when a.sex=0 or a.sex is null then 1 end) as '未知'
    from hopsonone_members.members_info a,
    hopsonone_members.members b,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and a.m_id=b.m_id
    and b.market_id= ma.id
    group by ma.id
    """
    # 计时
    累计会员量运行开始时间 = dt.datetime.now()
    累计会员量 = pd.read_sql(累计会员量sql, con=连接对象, params={'date_ed': 结束日期}, index_col='商场名称')
    累计会员量运行结束时间 = dt.datetime.now()
    累计会员量运行sql用时 = (累计会员量运行结束时间 - 累计会员量运行开始时间).seconds
    print(f'累计会员量运行SQL时间为{累计会员量运行sql用时}秒')

    # 本月会员新增SQL语句
    本月会员新增sql = """
    SELECT ma.market_name as '商场名称',
    sum(case when b.create_time>=DATE_SUB(%(date_op)s, INTERVAL 1 month)
    and b.create_time<=concat(DATE_SUB(%(date_op)s, INTERVAL 1 day),' 23:59:59') then 1 end) as '上月新增会员',
    sum(case when b.create_time>=%(date_op)s
    and b.create_time<=concat(%(date_ed)s,' 23:59:59') then 1 end) as '本月新增会员'
    from hopsonone_members.members b,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time>=DATE_SUB(%(date_op)s, INTERVAL 1 month)
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and b.market_id= ma.id
    group by ma.id
    limit 0, 5000
    """
    # 计时
    本月会员新增运行开始时间 = dt.datetime.now()
    本月会员新增 = pd.read_sql(本月会员新增sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    本月会员新增运行结束时间 = dt.datetime.now()
    本月会员新增运行sql用时 = (本月会员新增运行结束时间 - 本月会员新增运行开始时间).seconds
    print(f'本月会员新增运行SQL时间为{本月会员新增运行sql用时}秒')

    # 会员年龄段统计SQL语句
    会员年龄段统计sql = """
    SELECT ma.market_name as '商场名称',
    sum(case when year(a.birthday)<1970 then 1 else 0 end) as '70前',
    sum(case when year (a.birthday)<1980 then 1 else 0 end) as '70后',
    sum(case when year (a.birthday)<1990 then 1 else 0 end) as '80后',
    sum(case when year (a.birthday)<2000 then 1 else 0 end) as '90后',
    sum(case when year (a.birthday)>=2000 then 1 else 0 end) as '00后'
    from hopsonone_members.members_info a,
    hopsonone_members.members b,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and a.birthday is not null
    and b.market_id= ma.id
    and a.m_id=b.m_id
    group by ma.id
    limit 0, 5000
    """
    # 计时
    会员年龄段统计运行开始时间 = dt.datetime.now()
    会员年龄段统计 = pd.read_sql(会员年龄段统计sql, con=连接对象, params={'date_ed': 结束日期}, index_col='商场名称')
    会员年龄段统计运行结束时间 = dt.datetime.now()
    会员年龄段统计运行sql用时 = (会员年龄段统计运行结束时间 - 会员年龄段统计运行开始时间).seconds
    print(f'会员年龄段统计运行SQL时间为{会员年龄段统计运行sql用时}秒')

    # 会员生日SQL语句
    会员生日sql = """
    SELECT ma.market_name as '商场名称',
    sum(case when month(a.birthday)=month(%(date_op)s) then 1 else 0 end) as '本月生日',
    sum(case when month(a.birthday)=month(%(date_op)s)+1 then 1 else 0 end) as '次月生日'
    from hopsonone_members.members_info a,
    hopsonone_members.members b,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and a.birthday is not null
    and b.market_id= ma.id
    and a.m_id=b.m_id
    group by ma.id
    limit 0, 5000
    """
    # 计时
    会员生日运行开始时间 = dt.datetime.now()
    会员生日 = pd.read_sql(会员生日sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    会员生日运行结束时间 = dt.datetime.now()
    会员生日运行sql用时 = (会员生日运行结束时间 - 会员生日运行开始时间).seconds
    print(f'会员生日运行SQL时间为{会员生日运行sql用时}秒')

    # 会员注册渠道SQL语句
    会员注册渠道sql = """
    SELECT ma.market_name as '商场名称',s.dic_desc as '渠道',count(*) as '会员数'
    from hopsonone_members.members b,
    (SELECT  distinct s1.dic_value,s1.dic_desc
    from hopsonone_cms.system_sys_dic s1
    where s1.type_name='REGISTER_TYPE') s,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time>=%(date_op)s
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and b.m_origin=s.dic_value
    and ma.id=b.market_id
    group by ma.id,s.dic_desc	
    order by count(*) desc
    limit 0, 5000
    """
    # 计时
    会员注册渠道运行开始时间 = dt.datetime.now()
    会员注册渠道 = pd.read_sql(会员注册渠道sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    会员注册渠道运行结束时间 = dt.datetime.now()
    会员注册渠道运行sql用时 = (会员注册渠道运行结束时间 - 会员注册渠道运行开始时间).seconds
    print(f'会员注册渠道运行SQL时间为{会员注册渠道运行sql用时}秒')

    # 线下销售及笔数SQL语句
    线下销售及笔数sql = """
    select ma.market_name as '商场名称',
    sum(a.adjust_tax_sales_amount_month)/100 as '整体销售（元）',
    sum(a.trade_count) as '销售笔数'
    from shop_side_operation_real_time.sales_report_details a,
    merchant_entity.entity_store b,
    merchant_entity.market ma
    where ma.id=108
    and a.trade_date>=%(date_op)s
    and a.trade_date<=concat(%(date_ed)s,' 23:59:59') 
    and b.market_id = ma.id
    and a.business_id = b.store_id
    group by ma.id
    limit 0, 5000
    """
    # 计时
    线下销售及笔数运行开始时间 = dt.datetime.now()
    线下销售及笔数 = pd.read_sql(线下销售及笔数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    线下销售及笔数运行结束时间 = dt.datetime.now()
    线下销售及笔数运行sql用时 = (线下销售及笔数运行结束时间 - 线下销售及笔数运行开始时间).seconds
    print(f'线下销售及笔数运行SQL时间为{线下销售及笔数运行sql用时}秒')

    # 会员销售及笔数SQL语句
    会员销售及笔数sql = """
    select ma.market_name as '商场名称',
    round(sum(case when p.state = 0 then ifnull(p.consume_amount/100,0) 
    when p.state = 1 then -ifnull(p.consume_amount/100,0) 
    end ),2) as '会员销售额（元）',
    sum(case when p.state = 0 then 1 
    when p.state = 1 then -1 
    end ) as '会员交易笔数'
    from hopsonone_point_real_time.members_points_detail p,
    hopsonone_members.members m,
    merchant_entity.market ma
    where p.behavior in (8,20,24,26,27,28) 
    and ma.id=108
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59') 
    and m.m_status=1
    and p.m_id=m.m_id
    and p.market_id=ma.id
    group by ma.market_name

    """
    # 计时
    会员销售及笔数运行开始时间 = dt.datetime.now()
    会员销售及笔数 = pd.read_sql(会员销售及笔数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    会员销售及笔数运行结束时间 = dt.datetime.now()
    会员销售及笔数运行sql用时 = (会员销售及笔数运行结束时间 - 会员销售及笔数运行开始时间).seconds
    print(f'会员销售及笔数运行SQL时间为{会员销售及笔数运行sql用时}秒')

    # 会员消费人数SQL语句
    会员消费人数sql = """
    select ma.market_name as '商场名称',count(distinct p.m_id) as '会员消费人数'
    from hopsonone_point_real_time.members_points_detail p,
    hopsonone_members.members m,
    merchant_entity.market ma
    where p.state = 0 
    and p.behavior in (8,20,24,26,27,28) 
    and ma.id=108
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59') 
    and m.m_status=1
    and p.m_id=m.m_id
    and p.market_id=ma.id
    group by ma.market_name
    limit 0, 5000
    """
    # 计时
    会员消费人数运行开始时间 = dt.datetime.now()
    会员消费人数 = pd.read_sql(会员消费人数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    会员消费人数运行结束时间 = dt.datetime.now()
    会员消费人数运行sql用时 = (会员消费人数运行结束时间 - 会员消费人数运行开始时间).seconds
    print(f'会员消费人数运行SQL时间为{会员消费人数运行sql用时}秒')

    # PLUS会员销售及笔数SQL语句
    PLUS会员销售及笔数sql = """
    select ma.market_name as '商场名称',
    round(sum(case when p.state = 0 then ifnull(p.consume_amount/100,0) 
    when p.state = 1 then -ifnull(p.consume_amount/100,0) 
    end ),2) as 'PLUS会员销售额（元）',
    sum(case when p.state = 0 then 1 
    when p.state = 1 then -1 
    end ) as 'PLUS会员交易笔数'
    from hopsonone_point_real_time.members_points_detail p,
    hopsonone_personal.members_plusvip_log k,
    hopsonone_members.members m,
    merchant_entity.market ma
    where p.behavior in (8,20,24,26,27,28) 
    and ma.id=108
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59') 
    and k.due_time>=%(date_op)s
    and m.m_status in (0,1)
    and p.m_id=m.m_id
    and p.m_id=k.m_id
    and p.market_id=ma.id
    """
    # 计时
    PLUS会员销售及笔数运行开始时间 = dt.datetime.now()
    PLUS会员销售及笔数 = pd.read_sql(PLUS会员销售及笔数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    PLUS会员销售及笔数运行结束时间 = dt.datetime.now()
    PLUS会员销售及笔数运行sql用时 = (PLUS会员销售及笔数运行结束时间 - PLUS会员销售及笔数运行开始时间).seconds
    print(f'PLUS会员销售及笔数运行SQL时间为{PLUS会员销售及笔数运行sql用时}秒')

    # PLUS会员消费人数SQL语句
    PLUS会员消费人数sql = """
    select ma.market_name as '商场名称',count(distinct p.m_id) as 'PLUS会员消费人数'
    from hopsonone_point_real_time.members_points_detail p,
    hopsonone_personal.members_plusvip_log k,
    hopsonone_members.members m,
    merchant_entity.market ma
    where p.state = 0 
    and p.behavior=20 
    and ma.id=108
    and p.consume_time>=%(date_op)s
    and p.consume_time<=concat(%(date_ed)s,' 23:59:59') 
    and m.m_status in (0,1)
    and p.m_id=m.m_id
    and p.m_id=k.m_id
    and p.market_id=ma.id
    """
    # 计时
    PLUS会员消费人数运行开始时间 = dt.datetime.now()
    PLUS会员消费人数 = pd.read_sql(PLUS会员消费人数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    PLUS会员消费人数运行结束时间 = dt.datetime.now()
    PLUS会员消费人数运行sql用时 = (PLUS会员消费人数运行结束时间 - PLUS会员消费人数运行开始时间).seconds
    print(f'PLUS会员消费人数运行SQL时间为{PLUS会员消费人数运行sql用时}秒')

    # 积分变动SQL语句
    积分变动sql = """
    select b.`商场名称`,b.`本月新增`,b.`本月消耗`,c.`累计剩余`
    from
    (select ma.market_name as '商场名称',
    sum(case when a.state=0 then a.actual_point end) as '本月新增',
    sum(case when a.state=1 then a.actual_point end) as '本月消耗'
    from hopsonone_point_real_time.members_points_detail a,
    merchant_entity.market ma
    where ma.id=108
    and a.create_time>=%(date_op)s
    and a.create_time<=concat(%(date_ed)s,' 23:59:59') 
    and a.market_id=ma.id
    group by a.market_id) b
    left join
    (select ma.market_name as '商场名称',sum(p.points) as '累计剩余'
    from hopsonone_point.members_points p,
    hopsonone_members.members m,
    merchant_entity.market ma
    where ma.id=108
    and p.update_time<=concat(%(date_ed)s,' 23:59:59') 
    and p.m_id=m.m_id
    and m.market_id=ma.id
    group by ma.id) c
    on b.`商场名称`=c.`商场名称`
    group by b.`商场名称`
    limit 0, 5000
    """
    # 计时
    积分变动运行开始时间 = dt.datetime.now()
    积分变动 = pd.read_sql(积分变动sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    积分变动运行结束时间 = dt.datetime.now()
    积分变动运行sql用时 = (积分变动运行结束时间 - 积分变动运行开始时间).seconds
    print(f'积分变动运行SQL时间为{积分变动运行sql用时}秒')

    积分段会员数sql = """
    select ma.market_name as '商场名称',
    sum(case when p.points<500 then 1 end) as '500以下',
    sum(case when p.points>=500 and p.points<=10000 then 1 end) as '500-1万',
    sum(case when p.points>10000 and p.points<=50000 then 1 end) as '1万-5万',
    sum(case when p.points>50000 and p.points<=100000 then 1 end) as '5万-10万',
    sum(case when p.points>100000 then 1 end) as '10万以上'
    from hopsonone_point.members_points p,
    hopsonone_members.members m,
    merchant_entity.market ma
    where ma.id=108
    and p.update_time<=concat(%(date_ed)s,' 23:59:59') 
    and p.m_id=m.m_id
    and m.market_id=ma.id
    group by ma.id
    """
    # 计时
    积分段会员数运行开始时间 = dt.datetime.now()
    积分段会员数 = pd.read_sql(积分段会员数sql, con=连接对象, params={'date_ed': 结束日期}, index_col='商场名称')
    积分段会员数运行结束时间 = dt.datetime.now()
    积分段会员数运行sql用时 = (积分段会员数运行结束时间 - 积分段会员数运行开始时间).seconds
    print(f'积分段会员数运行SQL时间为{积分段会员数运行sql用时}秒')
    # 积分业态前三SQL语句
    积分业态前三sql = """
    select ma.market_name as '商场名称',t.`业态`
    from (SELECT p.market_id,s.dic_desc as '业态',sum(p.point_amount) as 'point_amount'
    from (select p.market_id,p.business_id,sum(p.consume_amount/100) as 'point_amount'
    from hopsonone_point_real_time.members_points_detail p
    where p.market_id=108
    and p.state = 0 
    and p.behavior in (8,20,24,26,27,28)
    and p.create_time>=%(date_op)s
    and p.create_time<=concat(%(date_ed)s,' 23:59:59')
    group by p.market_id,p.business_id) p,
    merchant_entity.entity_store bu,
    hopsonone_cms.sys_dic s
    where s.type_name = 'operationtype'
    and p.business_id=bu.store_id
    and bu.format_type = s.dic_value
    and bu.market_id = s.market_id
    group by p.market_id,s.dic_desc) t,
    merchant_entity.market ma
    where ma.id=t.market_id
    order by t.point_amount desc
    limit 3
    """
    # 计时
    积分业态前三运行开始时间 = dt.datetime.now()
    积分业态前三 = pd.read_sql(积分业态前三sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    积分业态前三运行结束时间 = dt.datetime.now()
    积分业态前三运行sql用时 = (积分业态前三运行结束时间 - 积分业态前三运行开始时间).seconds
    print(f'积分业态前三运行SQL时间为{积分业态前三运行sql用时}秒')

    # PLUS数据分析SQL语句
    PLUS数据分析sql = """
    select a1.market_name as '商场名称',
    a1.`累计购买量`,
    a1.`本月新增`,
    a1.`上月新增`,
    concat(round(ifnull((a1.`本月新增`-a1.`上月新增`)/a1.`上月新增`,0)*100,2),'%%') as '环比',
    concat(round(ifnull((a1.`本月新增`-a1.`同期`)/a1.`同期`,0)*100,2),'%%') as '同比',
    a1.`累计购买金额`,
    a1.`本月新增金额`,
    a1.`本月复购量`
    from
    (select ma.market_name,
    count((case when o.payment_dt>='2022-01-01'
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.buyer_id end)) as '累计购买量',
    count((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 month)
    and o.payment_dt<=concat(DATE_SUB(%(date_op)s, INTERVAL 1 day),' 23:59:59') then o.buyer_id end)) as '上月新增',
    count((case when o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
    and o.payment_dt<=concat(DATE_SUB(%(date_ed)s, INTERVAL 1 year),' 23:59:59') then o.buyer_id end)) as '同期',
    count((case when o.payment_dt>=%(date_op)s
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.buyer_id end)) as '本月新增',
    sum((case when o.payment_dt>='2022-01-01'
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.total_fee end))/100 as '累计购买金额',
    sum((case when o.payment_dt>=%(date_op)s
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') then o.total_fee end))/100 as '本月新增金额',
    count((case when o.payment_dt>=%(date_op)s
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59') and a.opt_type=2 then o.buyer_id end)) as '本月复购量'
    from hopsonone_order_center.orders_detail_plusvip a,
    merchant_entity.market ma,
    hopsonone_order_center.orders o
    where a.trade_no = o.trade_no
    and o.market_id in (218,108,237,110,132,213,164,234)
    and o.state in (3, 4, 5)
    and ma.id=108
    and o.market_id=ma.id
    and o.payment_dt>=DATE_SUB(%(date_op)s, INTERVAL 1 year)
    and o.payment_dt<=concat(%(date_ed)s,' 23:59:59')
    GROUP BY o.market_id) a1
    """
    # 计时
    PLUS数据分析运行开始时间 = dt.datetime.now()
    PLUS数据分析 = pd.read_sql(PLUS数据分析sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    PLUS数据分析运行结束时间 = dt.datetime.now()
    PLUS数据分析运行sql用时 = (PLUS数据分析运行结束时间 - PLUS数据分析运行开始时间).seconds
    print(f'PLUS数据分析运行SQL时间为{PLUS数据分析运行sql用时}秒')

    # 业态销售排名SQL语句
    业态销售排名sql = """
    SELECT ma.market_name as '商场名称',
    s.dic_desc as '业态名称',
    sum(a.adjust_tax_sales_amount_month)/100 as '销售金额',
    sum(a.trade_count) as '销售单数'
    FROM shop_side_operation_real_time.sales_report_details a, 
    merchant_entity.entity_store b,
    hopsonone_cms.sys_dic s,
    merchant_entity.market ma
    where ma.id=108
    and a.trade_date>=%(date_op)s
    and a.trade_date<=concat(%(date_ed)s,' 23:59:59') 
    and s.type_name = 'operationtype' 
    and b.market_id = ma.id
    and a.business_id = b.store_id
    and b.format_type = s.dic_value
    and b.market_id = s.market_id
    group by ma.id,s.dic_desc
    order by sum(a.adjust_tax_sales_amount_month) desc
    limit 0, 5000
    """
    # 计时
    业态销售排名运行开始时间 = dt.datetime.now()
    业态销售排名 = pd.read_sql(业态销售排名sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    业态销售排名运行结束时间 = dt.datetime.now()
    业态销售排名运行sql用时 = (业态销售排名运行结束时间 - 业态销售排名运行开始时间).seconds
    print(f'业态销售排名运行SQL时间为{业态销售排名运行sql用时}秒')

    # 停车券SQL语句
    停车券sql = """
    select a1.`商场名称`,a1.`卡券领取`,a2.`卡券核销`,concat(round(a2.`卡券核销`/a1.`卡券领取`*100,2),'%%') as '核销率'
    from
    (select ma.market_name as '商场名称',
    count(*) as '卡券领取'
    from hopsonone_coupons_v2.coupons_sub_order a,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and a.receive_time>=%(date_op)s
    and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
    and a.order_state=3 #已支付
    and a.type=0 #停车
    and c.use_scope=2
    and c.id=a.coupons_id
    and a.market_id=ma.id
    group by a.market_id) a1
    left join
    (select ma.market_name as '商场名称',
    count(*) as '卡券核销'
    from hopsonone_coupons_v2.coupons_use_record b,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and b.create_time>=%(date_op)s
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and c.use_scope =2
    and c.id=b.coupons_id
    and b.market_id=ma.id
    group by b.market_id) a2
    on a1.`商场名称`=a2.`商场名称`

    limit 0, 5000
    """
    # 计时
    停车券运行开始时间 = dt.datetime.now()
    停车券 = pd.read_sql(停车券sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    停车券运行结束时间 = dt.datetime.now()
    停车券运行sql用时 = (停车券运行结束时间 - 停车券运行开始时间).seconds
    print(f'停车券运行SQL时间为{停车券运行sql用时}秒')

    # 团购券SQL语句
    团购券sql = """
    select a1.`商场名称`,a1.`卡券领取`,a2.`卡券核销`,concat(round(a2.`卡券核销`/a1.`卡券领取`*100,2),'%%') as '核销率'
    from
    (select ma.market_name as '商场名称',
    count(*) as '卡券领取'
    from hopsonone_coupons_v2.coupons_sub_order a,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and a.receive_time>=%(date_op)s
    and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
    and a.order_state=3 #已支付
    and a.type !=0 #非团购
    and c.use_scope !=2
    and c.get_way in (2,3)
    and c.id=a.coupons_id
    and a.market_id=ma.id
    group by a.market_id) a1
    left join
    (select ma.market_name as '商场名称',
    count(*) as '卡券核销'
    from hopsonone_coupons_v2.coupons_use_record b,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and b.create_time>=%(date_op)s
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and c.use_scope !=2
    and c.get_way in (2,3)
    and c.id=b.coupons_id
    and b.market_id=ma.id
    group by b.market_id) a2
    on a1.`商场名称`=a2.`商场名称`
    limit 0, 5000
    """
    # 计时
    团购券运行开始时间 = dt.datetime.now()
    团购券 = pd.read_sql(团购券sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    团购券运行结束时间 = dt.datetime.now()
    团购券运行sql用时 = (团购券运行结束时间 - 团购券运行开始时间).seconds
    print(f'团购券运行SQL时间为{团购券运行sql用时}秒')

    # 体验券SQL语句
    体验券sql = """
    select a1.`商场名称`,a1.`卡券领取`,a2.`卡券核销`,concat(round(a2.`卡券核销`/a1.`卡券领取`*100,2),'%%') as '核销率'
    from
    (select ma.market_name as '商场名称',
    count(*) as '卡券领取'
    from hopsonone_coupons_v2.coupons_sub_order a,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and a.receive_time>=%(date_op)s
    and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
    and a.order_state=3 #已支付
    and a.type !=0 #非停车
    and c.use_scope !=2
    and c.get_way=1
    and c.id=a.coupons_id
    and a.market_id=ma.id
    group by a.market_id) a1
    left join
    (select ma.market_name as '商场名称',
    count(*) as '卡券核销'
    from hopsonone_coupons_v2.coupons_use_record b,
    merchant_entity.market ma
    where ma.id=108
    and b.create_time>=%(date_op)s
    and b.create_time<=concat(%(date_ed)s,' 23:59:59')
    and b.coupons_id in 
    (select distinct a.coupons_id
    from hopsonone_coupons_v2.coupons_sub_order a,
    merchant_entity.market ma,
    hopsonone_coupons_v2.coupons c
    where ma.id=108
    and a.receive_time>=%(date_op)s
    and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
    and a.order_state=3 #已支付
    and a.type !=0 #非停车
    and c.use_scope !=2
    and c.get_way=1
    and c.id=a.coupons_id)
    and b.market_id=ma.id
    group by b.market_id) a2
    on a1.`商场名称`=a2.`商场名称`

    limit 0, 5000
    """
    # 计时
    体验券运行开始时间 = dt.datetime.now()
    体验券 = pd.read_sql(体验券sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    体验券运行结束时间 = dt.datetime.now()
    体验券运行sql用时 = (体验券运行结束时间 - 体验券运行开始时间).seconds
    print(f'体验券运行SQL时间为{体验券运行sql用时}秒')

    # 活跃会员率SQL语句
    活跃会员率sql = """
    select
    case c.market_id
    when 108 then '成都温江合生汇'
    end as '商场名称',
    concat(round(c.`活跃会员`/b.`会员量`*100,2),'%%') as '活跃会员率'
    from (select a.market_id,count(*) as '活跃会员'
    from hopsonone_do.active_m_id本月 a
    where market_id=108
    group by market_id) c,
    (select a.market_id,count(*) as '会员量'
    from hopsonone_members.members a
    where a.market_id=108
    and a.create_time<=concat(%(date_ed)s,' 23:59:59')
    GROUP BY a.market_id) b
    where c.market_id=b.market_id
    """
    # 计时
    活跃会员率运行开始时间 = dt.datetime.now()
    活跃会员率 = pd.read_sql(活跃会员率sql, con=连接对象, params={'date_ed': 结束日期}, index_col='商场名称')
    活跃会员率运行结束时间 = dt.datetime.now()
    活跃会员率运行sql用时 = (活跃会员率运行结束时间 - 活跃会员率运行开始时间).seconds
    print(f'活跃会员率运行SQL时间为{活跃会员率运行sql用时}秒')

    # plus会员活跃率SQL语句
    plus会员活跃率sql = """
    select case x.market_id
    when 108 then '成都温江合生汇'
    end as '商场名称',
    concat(round(count(distinct x.mid)/m.`plus总会员人数`*100,2),'%%') as 'plus会员活跃率'
    from hopsonone_do.active_m_idVPI本月 x,
    (select m.opt_market,count(DISTINCT m.m_id) 'plus总会员人数'
    from hopsonone_personal.members_plusvip_log m
    where m.opt_market=108
    and m.due_time>=%(date_op)s) m
    where x.market_id=m.opt_market
    """
    # 计时
    plus会员活跃率运行开始时间 = dt.datetime.now()
    plus会员活跃率 = pd.read_sql(plus会员活跃率sql, con=连接对象, params={'date_op': 开始日期}, index_col='商场名称')
    plus会员活跃率运行结束时间 = dt.datetime.now()
    plus会员活跃率运行sql用时 = (plus会员活跃率运行结束时间 - plus会员活跃率运行开始时间).seconds
    print(f'plus会员活跃率运行SQL时间为{plus会员活跃率运行sql用时}秒')

    # 工作簿列表=项目月度总结.sheetnames
    会员数据 = 项目月度总结['会员数据']
    会员数据['C5'].value = 累计会员量.loc['成都温江合生汇']['男']
    会员数据['D5'].value = 累计会员量.loc['成都温江合生汇']['女']
    会员数据['E5'].value = 累计会员量.loc['成都温江合生汇']['未知']
    会员数据['G5'].value = 本月会员新增.loc['成都温江合生汇']['本月新增会员']
    会员数据['H5'].value = 本月会员新增.loc['成都温江合生汇']['上月新增会员']
    会员数据['J5'].value = 活跃会员率.loc['成都温江合生汇']['活跃会员率']
    会员数据['K5'].value = plus会员活跃率.loc['成都温江合生汇']['plus会员活跃率']
    会员数据['L5'].value = 会员年龄段统计.loc['成都温江合生汇']['70前']
    会员数据['M5'].value = 会员年龄段统计.loc['成都温江合生汇']['70后']
    会员数据['N5'].value = 会员年龄段统计.loc['成都温江合生汇']['80后']
    会员数据['O5'].value = 会员年龄段统计.loc['成都温江合生汇']['90后']
    会员数据['P5'].value = 会员年龄段统计.loc['成都温江合生汇']['00后']
    会员数据['Q5'].value = 会员生日.loc['成都温江合生汇']['本月生日']
    会员数据['R5'].value = 会员生日.loc['成都温江合生汇']['次月生日']
    for i in range(len(会员注册渠道.loc['成都温江合生汇']['渠道'])):
        会员数据.cell(4, 19 + i).value = 会员注册渠道.loc['成都温江合生汇']['渠道'][i]
        会员数据.cell(5, 19 + i).value = 会员注册渠道.loc['成都温江合生汇']['会员数'][i]

    积分消费数据分析 = 项目月度总结['积分消费数据分析']
    积分消费数据分析['C5'].value = 线下销售及笔数.loc['成都温江合生汇']['整体销售（元）']
    积分消费数据分析['D5'].value = 线下销售及笔数.loc['成都温江合生汇']['销售笔数']
    积分消费数据分析['E5'].value = 会员销售及笔数.loc['成都温江合生汇']['会员销售额（元）']
    积分消费数据分析['F5'].value = 会员消费人数.loc['成都温江合生汇']['会员消费人数']
    积分消费数据分析['G5'].value = 会员销售及笔数.loc['成都温江合生汇']['会员交易笔数']
    积分消费数据分析['I5'].value = PLUS会员销售及笔数.loc['成都温江合生汇']['PLUS会员销售额（元）']
    积分消费数据分析['J5'].value = PLUS会员消费人数.loc['成都温江合生汇']['PLUS会员消费人数']
    积分消费数据分析['K5'].value = PLUS会员销售及笔数.loc['成都温江合生汇']['PLUS会员交易笔数']
    积分消费数据分析['N5'].value = 积分变动.loc['成都温江合生汇']['本月新增']
    积分消费数据分析['O5'].value = 积分变动.loc['成都温江合生汇']['本月消耗']
    积分消费数据分析['P5'].value = 积分变动.loc['成都温江合生汇']['累计剩余']
    积分消费数据分析['Q5'].value = 积分段会员数.loc['成都温江合生汇']['500以下']
    积分消费数据分析['R5'].value = 积分段会员数.loc['成都温江合生汇']['500-1万']
    积分消费数据分析['S5'].value = 积分段会员数.loc['成都温江合生汇']['1万-5万']
    积分消费数据分析['T5'].value = 积分段会员数.loc['成都温江合生汇']['5万-10万']
    积分消费数据分析['U5'].value = 积分段会员数.loc['成都温江合生汇']['10万以上']
    积分消费数据分析['V5'].value = 积分业态前三.loc['成都温江合生汇']['业态'][0]
    积分消费数据分析['W5'].value = 积分业态前三.loc['成都温江合生汇']['业态'][1]
    积分消费数据分析['X5'].value = 积分业态前三.loc['成都温江合生汇']['业态'][2]

    PLUS数据 = 项目月度总结['PLUS数据']
    PLUS数据['C5'].value = PLUS数据分析.loc['成都温江合生汇']['累计购买量']
    PLUS数据['D5'].value = PLUS数据分析.loc['成都温江合生汇']['本月新增']
    PLUS数据['E5'].value = PLUS数据分析.loc['成都温江合生汇']['上月新增']
    PLUS数据['F5'].value = PLUS数据分析.loc['成都温江合生汇']['环比']
    PLUS数据['G5'].value = PLUS数据分析.loc['成都温江合生汇']['同比']
    PLUS数据['H5'].value = PLUS数据分析.loc['成都温江合生汇']['累计购买金额']
    PLUS数据['I5'].value = PLUS数据分析.loc['成都温江合生汇']['本月新增金额']
    PLUS数据['J5'].value = PLUS数据分析.loc['成都温江合生汇']['本月复购量']

    积分业态销售排名 = 项目月度总结['积分业态销售排名']
    for j in range(len(业态销售排名.loc['成都温江合生汇']['业态名称'])):
        积分业态销售排名.cell(2 + j, 1).value = 业态销售排名.loc['成都温江合生汇']['业态名称'][j]
        积分业态销售排名.cell(2 + j, 2).value = 业态销售排名.loc['成都温江合生汇']['销售金额'][j]
        积分业态销售排名.cell(2 + j, 3).value = 业态销售排名.loc['成都温江合生汇']['销售单数'][j]

    卡券核销 = 项目月度总结['卡券核销']
    卡券核销['B3'].value = 停车券.loc['成都温江合生汇']['卡券领取']
    卡券核销['C3'].value = 停车券.loc['成都温江合生汇']['卡券核销']
    卡券核销['D3'].value = 停车券.loc['成都温江合生汇']['核销率']
    卡券核销['E3'].value = 团购券.loc['成都温江合生汇']['卡券领取']
    卡券核销['F3'].value = 团购券.loc['成都温江合生汇']['卡券核销']
    卡券核销['G3'].value = 团购券.loc['成都温江合生汇']['核销率']
    卡券核销['H3'].value = 体验券.loc['成都温江合生汇']['卡券领取']
    卡券核销['I3'].value = 体验券.loc['成都温江合生汇']['卡券核销']
    卡券核销['J3'].value = 体验券.loc['成都温江合生汇']['核销率']
    项目月度总结.save(output_dir +'温江项目月度总结-'+周报日期+'.xlsx')
    print('完成')

if __name__ == '__main__':
    output_dir = './out/温江项目月度总结/'
    os.system('rm -rf ./out/温江项目月度总结/*')
    start_rq, end_rq = get_last_month_first_last_day()
    week_rq = start_rq[5:7] + start_rq[8:] + '~' + end_rq[5:7] + end_rq[8:]

    print('温江项目月度总结（月报）:')
    print('-----------------------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('周报日期：', week_rq)
    print('-----------------------------------------')

    # preprocessing
    exec_proc('proc_month_vip_member_active', start_rq, end_rq)
    exec_proc('proc_month_member_active',start_rq,end_rq)

    # write excel
    write_month_108(start_rq,end_rq,week_rq)

    # compressed files
    path_original = 'out/温江项目月度总结'
    path_zip = './out/温江项目月度总结/温江项目月度总结-{}.zip'.format(week_rq)
    zip_name = '温江项目月度总结-{}.zip'.format(week_rq)
    file_to_zip(path_original, path_zip)
    print('文件打包完成!')

    # send mail
    sender = '190634@lifeat.cn,820618@cre-hopson.com'
    Cc = '190343@lifeat.cn,820987@cre-hopson.com'
    title = '温江项目月度总结-{}'.format(week_rq)
    content = '''各位领导：
             {}
             详见附件,请查收。
             '''.format(title)
    os.chdir('./out/温江项目月度总结')
    file_list = [zip_name]
    ret = SendMail(sender, Cc, title, content).send(file_list)
    print(ret, type(ret))
    print('邮件邮件完成')