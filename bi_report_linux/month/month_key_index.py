#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/1/3 9:37
# @Author : ma.fei
# @File : month_key_index_218_110_108.py
# @Software: PyCharm

import pymysql
import pandas as pd
import datetime as dt
import os
import openpyxl
import zipfile
import pathlib
import smtplib
from datetime import datetime,timedelta
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

class SendMail(object):
    '''
    https://cloud.tencent.com/developer/article/1598257
    '''
    def __init__(self,sender,Cc,title,content):
        self.sender = sender.split(',')  #发送地址
        self.Cc = Cc.split(',')  # 抄送地址
        self.title = title  # 标题
        self.content = content  # 发送内容
        self.sys_sender = '190343@lifeat.cn'  # 系统账户
        self.sys_pwd = 'R86hyfjobMBYR76h'  # 系统账户密码

    def send(self,file_list):
        """
        发送邮件
        :param file_list: 附件文件列表
        :return: bool
        """
        try:
            # 创建一个带附件的实例
            msg = MIMEMultipart()
            # 发件人格式
            msg['From'] = formataddr(["技术服务部", self.sys_sender])
            # 收件人格式
            print('a=',[formataddr(["", u]) for u in self.sender])
            msg['To'] = sender

            # 抄送人
            msg['Cc'] = Cc

            # 邮件主题
            msg['Subject'] = self.title

            # 邮件正文内容
            msg.attach(MIMEText(self.content, 'plain', 'utf-8'))

            # 多个附件
            for file_name in file_list:
                print("file_name",file_name)
                # 构造附件
                xlsxpart = MIMEApplication(open(file_name, 'rb').read())
                # filename表示邮件中显示的附件名
                xlsxpart.add_header('Content-Disposition','attachment',filename = '%s'%file_name)
                msg.attach(xlsxpart)

            # SMTP服务器
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465,timeout=10)
            # 登录账户
            server.login(self.sys_sender, self.sys_pwd)
            # 发送邮件
            #server.sendmail(self.sys_sender, [self.sender, ], msg.as_string())
            server.sendmail(self.sys_sender,
                            [formataddr(["", u]) for u in self.sender]+
                            [formataddr(["", u]) for u in self.Cc],
                            msg.as_string())
            # 退出账户
            server.quit()
            return True
        except Exception as e:
            print(e)
            return False

def file_to_zip(path_original, path_zip):
    '''
     作用：压缩文件到指定压缩包里
     参数一：压缩文件的位置
     参数二：压缩后的压缩包
    '''
    # 提前读取，避免把压缩包自己加上去
    # 这里用list()做一个克隆提前执行下，不然会在后面循环时才执行这一引用，如果压缩包在这个路径下，会将它读取进来。
    f_list = list(pathlib.Path(path_original).glob("**/*"))
    z = zipfile.ZipFile(path_zip, 'w')
    #print('len(path_original)=',len(path_original))
    for f in f_list:
        #print('f=',f,str(f),str(f)[len(path_original):])
        z.write(f, str(f)[len(path_original):])
    z.close()

def get_last_month_first_last_day():
    now = datetime.now()
    this_month_start = datetime(now.year, now.month, 1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = datetime(last_month_end.year, last_month_end.month, 1)
    v_last_month_start = datetime.strftime(last_month_start, "%Y-%m-%d")
    v_last_month_end = datetime.strftime(last_month_end, "%Y-%m-%d")
    return v_last_month_start, v_last_month_end

def format_sql(v_sql):
    return v_sql.replace("\\","\\\\").replace("'","\\'")

def get_ds_mysql(ip,port,service ,user,password):
    conn = pymysql.connect(host=ip,
                           port=int(port),
                           user=user,
                           passwd=password,
                           db=service,
                           charset='utf8',
                           cursorclass = pymysql.cursors.DictCursor,
                           autocommit=False)
    return conn

def exec_proc(p_proc_name,p_rq_begin,p_rq_end):
    db_mysql = get_ds_mysql('rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com','3306','hopsonone_do','hopsonone_do','8Loxk2IJxaenJkE3')
    cr_mysql = db_mysql.cursor()
    print('call {}...'.format(p_proc_name))
    cr_mysql.callproc(p_proc_name,args=(p_rq_begin,p_rq_end))
    print('call {}...complete'.format(p_proc_name))

def write_key_index(p_rq_start,p_rq_end,p_week_rq):
    开始日期, 结束日期, 周报日期 = p_rq_start, p_rq_end, p_week_rq

    # 取读模板
    项目关键指标月报 = openpyxl.load_workbook('month_key_index.xlsx')
    项目名称列表 = 项目关键指标月报.sheetnames

    # 连接数据库，执行SQL语句
    连接对象 = pymysql.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',
                      passwd='8Loxk2IJxaenJkE3', database='hopsonone_do')
    # 会员注册量SQL语句
    当月会员注册量sql = """
    select ma.market_name as '商场名称',count(a.m_id) as '会员量'
    from hopsonone_members.members a,
    merchant_entity.market ma
    where a.market_id in (110,108,218)
    and a.create_time>=%(date_op)s
    and a.create_time<=concat(%(date_ed)s,' 23:59:59')
    and a.market_id=ma.id
    GROUP BY a.market_id
    """
    # 计时
    当月会员注册量运行开始时间 = dt.datetime.now()
    当月会员注册量 = pd.read_sql(当月会员注册量sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    当月会员注册量运行结束时间 = dt.datetime.now()
    当月会员注册量运行sql用时 = (当月会员注册量运行结束时间 - 当月会员注册量运行开始时间).seconds
    print(f'当月会员注册量运行SQL时间为{当月会员注册量运行sql用时}秒')

    # 总会员数SQL语句
    总会员数sql = """
    select ma.market_name as '商场名称',count(*) '总会员人数'
    from hopsonone_members.members m,
    merchant_entity.market ma
    where m.market_id in (110,108,218)
    and m.create_time<=concat(%(date_ed)s,' 23:59:59')
    and m.market_id=ma.id
    group by m.market_id
    """
    # 计时
    总会员数运行开始时间 = dt.datetime.now()
    总会员数 = pd.read_sql(总会员数sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    总会员数运行结束时间 = dt.datetime.now()
    总会员数运行sql用时 = (总会员数运行结束时间 - 总会员数运行开始时间).seconds
    print(f'总会员数运行SQL时间为{总会员数运行sql用时}秒')

    # 线下销售额SQL语句
    线下销售额sql = """
    SELECT ma.market_name as '商场名称',
    round(sum(a.adjust_tax_sales_amount_month)/100/10000,2) as '线下销售额'
    FROM shop_side_operation_real_time.sales_report_details a, 
    merchant_entity.entity_store b,
    merchant_entity.market ma
    WHERE a.business_id = b.store_id
    and b.market_id in (110,108,218)
    and b.market_id=ma.id
    and a.trade_date>=%(date_op)s
    and a.trade_date<=%(date_ed)s
    group by b.market_id
    """
    # 计时
    线下销售额运行开始时间 = dt.datetime.now()
    线下销售额 = pd.read_sql(线下销售额sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    线下销售额运行结束时间 = dt.datetime.now()
    线下销售额运行sql用时 = (线下销售额运行结束时间 - 线下销售额运行开始时间).seconds
    print(f'线下销售额运行SQL时间为{线下销售额运行sql用时}秒')

    # 线下会员销售SQL语句
    线下会员销售sql = """
    SELECT ma.market_name  as '商场名称',
    round(sum(case when p.state = 0 then ifnull(p.consume_amount/100,0) 
    when p.state = 1 then -ifnull(p.consume_amount/100,0) 
    end )/10000,2) as '线下会员消费'
    FROM hopsonone_point_real_time.members_points_detail p,
    merchant_entity.market ma
    where p.market_id=ma.id
    and p.market_id in (218,110,108)
    and p.behavior in (8,20,24,26,27,28)
    and p.consume_time >=%(date_op)s
    and p.consume_time <=concat(%(date_ed)s,' 23:59:59')
    group by p.market_id
    """
    # 计时
    线下会员销售运行开始时间 = dt.datetime.now()
    线下会员销售 = pd.read_sql(线下会员销售sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    线下会员销售运行结束时间 = dt.datetime.now()
    线下会员销售运行sql用时 = (线下会员销售运行结束时间 - 线下会员销售运行开始时间).seconds
    print(f'线下会员销售运行SQL时间为{线下会员销售运行sql用时}秒')

    # 付费会员复购SQL语句
    付费会员复购sql = """
    select ma.market_name as '商场名称',
    count(distinct m_id)as '付费会员复购会员数'
    from hopsonone_personal.members_plusvip_log p,
    merchant_entity.market ma
    where p.status=2
    and p.opt_market in (108,110,218)
    and p.create_time>='2021-01-01'
    and p.create_time<=concat(%(date_ed)s,' 23:59:59')
    and p.opt_type=2
    and p.opt_market=ma.id
    group by p.opt_market
    """
    # 计时
    付费会员复购运行开始时间 = dt.datetime.now()
    付费会员复购 = pd.read_sql(付费会员复购sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    付费会员复购运行结束时间 = dt.datetime.now()
    付费会员复购运行sql用时 = (付费会员复购运行结束时间 - 付费会员复购运行开始时间).seconds
    print(f'付费会员复购运行SQL时间为{付费会员复购运行sql用时}秒')

    # 付费会员累计SQL语句
    付费会员累计sql = """
    select ma.market_name as '商场名称',
    count(distinct m_id) as '付费会员累计人数'
    from hopsonone_personal.members_plusvip_log p,
    merchant_entity.market ma
    where p.status=2
    and p.opt_market in (108,110,218)
    and p.create_time<=concat(%(date_ed)s,' 23:59:59')
    and p.opt_market=ma.id
    group by p.opt_market
    """
    # 计时
    付费会员累计运行开始时间 = dt.datetime.now()
    付费会员累计 = pd.read_sql(付费会员累计sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    付费会员累计运行结束时间 = dt.datetime.now()
    付费会员累计运行sql用时 = (付费会员累计运行结束时间 - 付费会员累计运行开始时间).seconds
    print(f'付费会员累计运行SQL时间为{付费会员累计运行sql用时}秒')

    # 卡券领取SQL语句
    卡券领取sql = """
    select ma.market_name as '商场名称',
    count(a.trade_no) as '卡券领取'
    from hopsonone_coupons_v2.coupons_sub_order a,
    merchant_entity.market ma
    where a.market_id in (108,110,218)
    and a.market_id=ma.id
    and a.receive_time>=%(date_op)s
    and a.receive_time<=concat(%(date_ed)s,' 23:59:59')
    and a.order_state in (3,4,5)
    group by a.market_id
    """
    # 计时
    卡券领取运行开始时间 = dt.datetime.now()
    卡券领取 = pd.read_sql(卡券领取sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    卡券领取运行结束时间 = dt.datetime.now()
    卡券领取运行sql用时 = (卡券领取运行结束时间 - 卡券领取运行开始时间).seconds
    print(f'卡券领取运行SQL时间为{卡券领取运行sql用时}秒')

    # 卡券核销SQL语句
    卡券核销sql = """
    select ma.market_name as '商场名称',
    count(a.sub_trade_no) as '卡券核销'
    from hopsonone_coupons_v2.coupons_use_record a,
    merchant_entity.market ma
    where a.market_id in (108,110,218)
    and a.market_id=ma.id
    and a.create_time>=%(date_op)s
    and a.create_time<=concat(%(date_ed)s,' 23:59:59')
    group by a.market_id
    """
    # 计时
    卡券核销运行开始时间 = dt.datetime.now()
    卡券核销 = pd.read_sql(卡券核销sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    卡券核销运行结束时间 = dt.datetime.now()
    卡券核销运行sql用时 = (卡券核销运行结束时间 - 卡券核销运行开始时间).seconds
    print(f'卡券核销运行SQL时间为{卡券核销运行sql用时}秒')

    # 付费会员购买渠道SQL语句
    付费会员购买渠道sql = """
    select ma.market_name as '商场名称',
    (case p.channel when 1 then 'APP' when 2 then '微信' when 3 then '支付宝' else '其他' end) as '购买渠道',
    count(*) as '购买量'
    from hopsonone_personal.members_plusvip_log p,
    merchant_entity.market ma
    where p.status=2
    and p.opt_market in (108,110,218)
    and p.create_time>=%(date_op)s
    and p.create_time<=concat(%(date_ed)s,' 23:59:59')
    and p.opt_market=ma.id
    group by p.opt_market,p.channel
    """
    # 计时
    付费会员购买渠道运行开始时间 = dt.datetime.now()
    付费会员购买渠道 = pd.read_sql(付费会员购买渠道sql, con=连接对象, params={'date_op': 开始日期, 'date_ed': 结束日期}, index_col='商场名称')
    付费会员购买渠道运行结束时间 = dt.datetime.now()
    付费会员购买渠道运行sql用时 = (付费会员购买渠道运行结束时间 - 付费会员购买渠道运行开始时间).seconds
    print(f'付费会员购买渠道运行SQL时间为{付费会员购买渠道运行sql用时}秒')

    # 活跃会员SQL语句
    活跃会员sql = """
    select
    case market_id
    when 108 then '成都温江合生汇'
    when 110 then '上海五角场合生汇'
    when 218 then '北京朝阳合生汇'
    end as '商场名称'
    ,count(*) as '活跃会员'
    from hopsonone_do.active_m_id本月
    group by market_id
    """
    # 计时
    活跃会员运行开始时间 = dt.datetime.now()
    活跃会员 = pd.read_sql(活跃会员sql, con=连接对象, index_col='商场名称')
    活跃会员运行结束时间 = dt.datetime.now()
    活跃会员运行sql用时 = (活跃会员运行结束时间 - 活跃会员运行开始时间).seconds
    print(f'活跃会员运行SQL时间为{活跃会员运行sql用时}秒')

    for 项目名称 in 项目名称列表:
        数据表 = 项目关键指标月报[项目名称]
        数据表['B2'].value = 当月会员注册量.loc[项目名称]['会员量']
        数据表['B4'].value = 线下会员销售.loc[项目名称]['线下会员消费']
        数据表['B6'].value = 活跃会员.loc[项目名称]['活跃会员']
        数据表['C4'].value = 线下销售额.loc[项目名称]['线下销售额']
        数据表['C6'].value = 总会员数.loc[项目名称]['总会员人数']
        数据表['B13'].value = 付费会员复购.loc[项目名称]['付费会员复购会员数']
        数据表['C13'].value = 付费会员累计.loc[项目名称]['付费会员累计人数']
        数据表['B15'].value = 卡券领取.loc[项目名称]['卡券领取']
        数据表['C15'].value = 卡券核销.loc[项目名称]['卡券核销']
        for i in range(len(付费会员购买渠道.loc[项目名称]['购买渠道'])):
            数据表.cell(8 + i, 2).value = 付费会员购买渠道.loc[项目名称]['购买渠道'][i]
            数据表.cell(8 + i, 3).value = 付费会员购买渠道.loc[项目名称]['购买量'][i]
    项目关键指标月报.save(output_dir+'朝合、温江、五角场项目关键指标月报-'+周报日期+'.xlsx')
    print('完成')

if __name__ == '__main__':
    output_dir = './out/朝合、温江、五角场项目关键指标月报/'
    os.system('rm -rf ./out/朝合、温江、五角场项目关键指标月报/*')
    start_rq, end_rq = get_last_month_first_last_day()
    week_rq = start_rq[5:7] + start_rq[8:] + '~' + end_rq[5:7] + end_rq[8:]

    print('朝合、温江、五角场项目关键指标月报:')
    print('-----------------------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('周报日期：', week_rq)
    print('-----------------------------------------')

    # preprocessing
    exec_proc('proc_month_key_index',start_rq,end_rq)

    # write excel
    write_key_index(start_rq,end_rq,week_rq)

    # compressed files
    path_original = 'out/朝合、温江、五角场项目关键指标月报'
    path_zip = './out/朝合、温江、五角场项目关键指标月报/朝合、温江、五角场项目关键指标月报-{}.zip'.format(week_rq)
    zip_name = '朝合、温江、五角场项目关键指标月报-{}.zip'.format(week_rq)
    file_to_zip(path_original, path_zip)
    print('文件打包完成!')

    # send mail
    sender = '190634@lifeat.cn,850646@cre-hopson.com,820618@cre-hopson.com,546564@hopson.com.cn'
    Cc = '190343@lifeat.cn,820987@cre-hopson.com'
    title = '朝合、温江、五角场项目关键指标月报-{}'.format(week_rq)
    content = '''各位领导：
             {}
             详见附件,请查收。
             '''.format(title)
    # 附件列表
    os.chdir('./out/朝合、温江、五角场项目关键指标月报')
    file_list = [zip_name]
    ret = SendMail(sender, Cc, title, content).send(file_list)
    print(ret, type(ret))
    print('邮件邮件完成')
