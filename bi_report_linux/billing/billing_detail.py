#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os
import pymysql as py
import pandas as pd
import openpyxl
import warnings

def write_billing_total(p_rq_start,p_rq_end):
    conn = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',passwd='8Loxk2IJxaenJkE3', database='shop_bill')
    billing_st_total   = """SELECT 
    	m.id AS '项目ID',
    	m.market_name AS '项目名称',
    	CONCAT(%(date_op)s,'~',%(date_ed)s) AS '时间维度',
    	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0)+ IFNULL(SUM(b.`电子签章费用`),0)AS '总费用',
        IFNULL(SUM(b.`本月发布账单数`),0) AS '发布账单数',
    	IFNULL(SUM(b.`电子签章费用`),0) AS '电子签章费用',
    	IFNULL(SUM(a.`B2B缴费金额`+a.`B2C缴费金额`+a.`扫码付缴费金额`),0) AS '缴费总金额',
    	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0) AS '缴费总手续费',   
    	IFNULL(SUM(a.`B2B缴费金额`),0) AS 'B2B缴费金额',
    	IFNULL(SUM(a.`B2B缴费笔数`),0) AS 'B2B缴费笔数',
    	IFNULL(SUM(a.`B2B缴费笔数`*5),0) AS 'B2B缴费手续费',
    	IFNULL(SUM(a.`B2C缴费金额`),0) AS 'B2C缴费金额',
    	IFNULL(SUM(a.`B2C缴费手续费`),0) AS 'B2C缴费手续费',
    	IFNULL(SUM(a.`扫码付缴费金额`),0) AS '扫码付缴费金额',
    	IFNULL(SUM(a.`扫码付缴费手续费`),0) AS '扫码付缴费手续费'
    FROM merchant_entity.market m 
     LEFT JOIN 
        (SELECT o.market_id,
    	ROUND(SUM(CASE WHEN o.tran_type='0002' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2B缴费金额',
    	SUM(CASE WHEN o.tran_type='0002' THEN 1 ELSE 0 END) AS 'B2B缴费笔数',
    	ROUND(SUM(CASE WHEN o.tran_type='0001' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2C缴费金额',
    	SUM(CASE WHEN o.tran_type='0001' THEN ROUND(o.total_fee*0.002/100,2) ELSE 0 END) AS 'B2C缴费手续费',
    	ROUND(SUM(CASE WHEN o.tran_type='0009' THEN o.total_fee ELSE 0 END)/100,2) AS '扫码付缴费金额',
    	SUM(CASE WHEN o.tran_type='0009' THEN o.handling_fee/100 ELSE 0 END) AS '扫码付缴费手续费'
    FROM hft_business_pay.bill_order o
    WHERE o.complete_date>=%(date_op)s
    	AND o.complete_date<=concat(%(date_ed)s,' 23:59:59')
    	AND o.state=3	
    GROUP BY o.market_id) a  ON  m.id = a.market_id
    LEFT JOIN (
    SELECT b.`marketId`,
           COUNT(0) AS '本月发布账单数',
           COUNT(0)*0.4 AS '电子签章费用'
    FROM bill b
    WHERE  b.`beginDT`>=%(date_op)s
       AND b.`beginDT`<=concat(%(date_ed)s,' 23:59:59')
     GROUP BY b.`marketId`
    ) b ON m.id=b.`marketId`
    WHERE m.id IN(218,213,164,108,323,287,234,132,170,278,306,237,315,327,
                  150,155,159,183,188,194,203,249,252,277,1101,2181)
    GROUP BY m.`id`
    ORDER BY m.market_type,
    CASE 
    WHEN m.id =218 THEN 1
    WHEN m.id =213 THEN 2
    WHEN m.id =164 THEN 3
    WHEN m.id =108 THEN 4
    WHEN m.id =323 THEN 5
    WHEN m.id =287 THEN 6
    WHEN m.id =110 THEN 7
    WHEN m.id =234 THEN 8
    WHEN m.id =132 THEN 9
    WHEN m.id =170 THEN 10
    WHEN m.id =278 THEN 11
    WHEN m.id =306 THEN 12
    WHEN m.id =237 THEN 13
    WHEN m.id =315 THEN 14
    WHEN m.id =327 THEN 15
    WHEN m.id =150 THEN 16
    WHEN m.id =155 THEN 17
    WHEN m.id =159 THEN 18
    WHEN m.id =183 THEN 19
    WHEN m.id =188 THEN 20
    WHEN m.id =194 THEN 21
    WHEN m.id =203 THEN 22
    WHEN m.id =249 THEN 23
    WHEN m.id =252 THEN 24
    WHEN m.id =277 THEN 25
    WHEN m.id =1101 THEN 26
    WHEN m.id =2181 THEN 27
    ELSE 999 END"""
    billing_st_detail  = """SELECT 
    o.out_trade_no AS '交易流水号',
    CONCAT(o.complete_date," ",o.complete_time) AS '缴费时间',
    o.total_fee/100 '缴费金额',
    CASE o.tran_type
    WHEN '0002' THEN 5 
    WHEN '0001' THEN ROUND(o.total_fee*0.002/100,2) 
    ELSE o.handling_fee/100 END AS '手续费',
    CASE o.tran_type
    WHEN 0001 THEN '个人网银支付'
    WHEN 0002 THEN '企业网银支付' 
    WHEN 0009 THEN '扫码付支付' 
    END AS '支付方式',
    m.market_name AS '项目名称',
    o.hst_business_id AS '商家编码(合生通)',
    IFNULL(b.store_name,'') AS '商家名称(合生通)',
    o.shop_berth AS '铺位号',
    CASE o.pay_type
    WHEN 1 THEN '预缴费'
    WHEN 0 THEN '账单缴费'
    END AS '缴费类型'
    FROM hft_business_pay.bill_order o
    LEFT JOIN merchant_entity.entity_store b ON o.hst_business_id=b.store_id
    LEFT JOIN merchant_entity.market m ON o.market_id=m.id
    WHERE o.complete_date>=%(date_op)s
    AND o.complete_date<=concat(%(date_ed)s,' 23:59:59')
    AND o.market_id=%(market_id)s
    AND o.state=3"""
    billing_st_digital = """SELECT 
	m.`market_name`    AS '项目名称',
	b.`invoiceCode`    AS '缴款单号',
	b.`beginDT`        AS '计费日期',
	b.`paymentDT`      AS '最晚付款日期',
	b.`nTtotalAmount`  AS '不含税金额',
	b.`totalAmount`    AS '含税金额',
	b.`hstBusinessId`  AS '商铺编码（合生通）',
	(SELECT store_name 
	 FROM `merchant_entity`.`entity_store` s 
	   WHERE s.store_id = b.`hstBusinessId` )   AS '商家名称（合生通）',
	b.`businessCode`  AS '商铺编码',
	b.`brand`         AS '品牌名称',
	b.`shopBerth`     AS '铺位号',
	b.`shopName`      AS '单位名称'
FROM shop_bill.bill b,merchant_entity.market m
WHERE  m.id = b.`marketId`
   AND b.`beginDT`>=%(date_op)s
   AND b.`beginDT`<=concat(%(date_ed)s,' 23:59:59')
   AND b.`marketId`=%(market_id)s"""
    billing_df_total   = pd.read_sql(billing_st_total,
                                     con=conn,
                                     params={'date_op': p_rq_start, 'date_ed': p_rq_end},
                                     index_col='项目ID')
    for index,row in billing_df_total.iterrows():
        billing_wb = openpyxl.load_workbook('./billing_detail.xlsx')
        print('生成:【'+row['项目名称']+'】项目数字账单手续费数据...')
        # 写【对账单】sheet页
        billing_st_dzd = billing_wb['对账单']
        billing_st_dzd.cell(3, 1).value = row['项目名称']
        billing_st_dzd.cell(3, 2).value = row['时间维度']
        billing_st_dzd.cell(3, 3).value = row['总费用']
        billing_st_dzd.cell(3, 4).value = row['发布账单数']
        billing_st_dzd.cell(3, 5).value = row['电子签章费用']
        billing_st_dzd.cell(3, 6).value = row['缴费总金额']
        billing_st_dzd.cell(3, 7).value = row['缴费总手续费']
        billing_st_dzd.cell(3, 8).value = row['B2B缴费金额']
        billing_st_dzd.cell(3, 9).value = row['B2B缴费笔数']
        billing_st_dzd.cell(3, 10).value = row['B2B缴费手续费']
        billing_st_dzd.cell(3, 11).value = row['B2C缴费金额']
        billing_st_dzd.cell(3, 12).value = row['B2C缴费手续费']
        billing_st_dzd.cell(3, 13).value = row['扫码付缴费金额']
        billing_st_dzd.cell(3, 14).value = row['扫码付缴费手续费']
        # 写【缴费明细】sheet页
        billing_st_jfmx = billing_wb['缴费明细']
        #print('param:',{'market_id':index,'date_op': p_rq_start, 'date_ed': p_rq_end})
        billing_df_detail = pd.read_sql(billing_st_detail,con=conn,params={'market_id':index,'date_op': p_rq_start, 'date_ed': p_rq_end})
        r=2
        for index2, row2 in billing_df_detail.iterrows():
            billing_st_jfmx.cell(r,1).value = row2['交易流水号']
            billing_st_jfmx.cell(r,2).value = row2['缴费时间']
            billing_st_jfmx.cell(r,3).value = row2['缴费金额']
            billing_st_jfmx.cell(r,4).value = row2['手续费']
            billing_st_jfmx.cell(r,5).value = row2['支付方式']
            billing_st_jfmx.cell(r,6).value = row2['项目名称']
            billing_st_jfmx.cell(r,7).value = row2['商家编码(合生通)']
            billing_st_jfmx.cell(r,8).value = row2['商家名称(合生通)']
            billing_st_jfmx.cell(r,9).value = row2['铺位号']
            billing_st_jfmx.cell(r,10).value = row2['缴费类型']
            r = r + 1
        # 写【电子签章明细】sheet页
        billing_st_dzqz = billing_wb['电子签章明细']
        #print('param:', {'market_id': index, 'date_op': p_rq_start, 'date_ed': p_rq_end})
        billing_df_digital = pd.read_sql(billing_st_digital, con=conn,params={'market_id': index, 'date_op': p_rq_start, 'date_ed': p_rq_end})
        r = 2
        for index2, row2 in billing_df_digital.iterrows():
            billing_st_dzqz.cell(r, 1).value = row2['项目名称']
            billing_st_dzqz.cell(r, 2).value = row2['缴款单号']
            billing_st_dzqz.cell(r, 3).value = row2['计费日期']
            billing_st_dzqz.cell(r, 4).value = row2['最晚付款日期']
            billing_st_dzqz.cell(r, 5).value = row2['不含税金额']
            billing_st_dzqz.cell(r, 6).value = row2['含税金额']
            billing_st_dzqz.cell(r, 7).value = row2['商铺编码（合生通）']
            billing_st_dzqz.cell(r, 8).value = row2['商家名称（合生通）']
            billing_st_dzqz.cell(r, 9).value = row2['商铺编码']
            billing_st_dzqz.cell(r, 10).value = row2['品牌名称']
            billing_st_dzqz.cell(r, 11).value = row2['铺位号']
            billing_st_dzqz.cell(r, 12).value = row2['单位名称']
            r = r + 1
        #billing_wb.save('./out/2024年'+row['项目名称']+'项目1季度数字账单手续费.xlsx')
        billing_wb.save('./out/{}年{}项目{}-{}月数字账单手续费.xlsx'.
                        format(p_rq_start[0:4],
                               row['项目名称'],
                               str(int(p_rq_start[5:7])),
                               str(int(p_rq_end[5:7]))))

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    output_dir = './out/'
    start_rq = '2024-04-01'
    end_rq = '2024-06-30'
    print('合生通项目组-数字账单手续费项目清单:')
    print('-----------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('-----------------------------')
    write_billing_total(start_rq,end_rq)