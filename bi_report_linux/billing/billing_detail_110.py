#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os
import pymysql as py
import pandas as pd
import openpyxl
import warnings

def write_billing_total(p_rq_start,p_rq_end):
    conn = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',passwd='8Loxk2IJxaenJkE3', database='shop_bill')
    billing_st_total   = """SELECT 
	110 AS '项目ID',
	a.`公司名称` AS '项目名称',
	CONCAT(%(date_op)s,'~',%(date_ed)s) AS '时间维度',
	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0)+ IFNULL(SUM(b.`电子签章费用`),0)AS '总费用',
        IFNULL(SUM(b.`本月发布账单数`),0) AS '发布账单数',
	IFNULL(SUM(b.`电子签章费用`),0) AS '电子签章费用',
	IFNULL(SUM(a.`B2B缴费金额`+a.`B2C缴费金额`+a.`扫码付缴费金额`),0) AS '缴费总金额',
	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0) AS '缴费总手续费',   
	IFNULL(SUM(a.`B2B缴费金额`),0) AS 'B2B缴费金额',
	IFNULL(SUM(a.`B2B缴费笔数`),0) AS 'B2B缴费笔数',
	IFNULL(SUM(a.`B2B缴费笔数`*5),0) AS 'B2B缴费手续费',
	IFNULL(SUM(a.`B2C缴费金额`),0) AS 'B2C缴费金额',
	IFNULL(SUM(a.`B2C缴费手续费`),0) AS 'B2C缴费手续费',
	IFNULL(SUM(a.`扫码付缴费金额`),0) AS '扫码付缴费金额',
	IFNULL(SUM(a.`扫码付缴费手续费`),0) AS '扫码付缴费手续费'
FROM (SELECT 
        CASE o.btob_shop_no WHEN '651312105270019' THEN '上海大展投资管理有限公司'
                            WHEN '651312105270016' THEN '中先国际控股有限公司' END AS '公司名称',
	ROUND(SUM(CASE WHEN o.tran_type='0002' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2B缴费金额',
	SUM(CASE WHEN o.tran_type='0002' THEN 1 ELSE 0 END) AS 'B2B缴费笔数',
	ROUND(SUM(CASE WHEN o.tran_type='0001' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2C缴费金额',
	SUM(CASE WHEN o.tran_type='0001' THEN ROUND(o.total_fee*0.002/100,2) ELSE 0 END) AS 'B2C缴费手续费',
	ROUND(SUM(CASE WHEN o.tran_type='0009' THEN o.total_fee ELSE 0 END)/100,2) AS '扫码付缴费金额',
	SUM(CASE WHEN o.tran_type='0009' THEN o.handling_fee/100 ELSE 0 END) AS '扫码付缴费手续费'
FROM hft_business_pay.bill_order o
WHERE o.complete_date>=%(date_op)s
	AND o.complete_date<=concat(%(date_ed)s,' 23:59:59')
	AND o.state=3
	AND o.market_id=110	
GROUP BY 1) a
LEFT JOIN
(SELECT 
       b.`payeeCompany` AS '公司名称',
       COUNT(0) AS '本月发布账单数',
       COUNT(0)*0.4 AS '电子签章费用'
FROM shop_bill.bill b
WHERE  b.`beginDT`>=%(date_op)s
   AND b.`beginDT`<=concat(%(date_ed)s,' 23:59:59')
   AND b.`marketId`=110
 GROUP BY 1
) b ON a.`公司名称`=b.`公司名称`
GROUP BY a.`公司名称`"""
    billing_st_detail  = """SELECT 
    o.out_trade_no AS '交易流水号',
    CONCAT(o.complete_date," ",o.complete_time) AS '缴费时间',
    o.total_fee/100 '缴费金额',
    CASE o.tran_type
    WHEN '0002' THEN 5 
    WHEN '0001' THEN ROUND(o.total_fee*0.002/100,2) 
    ELSE o.handling_fee/100 END AS '手续费',
    CASE o.tran_type
    WHEN 0001 THEN '个人网银支付'
    WHEN 0002 THEN '企业网银支付' 
    WHEN 0009 THEN '扫码付支付' 
    END AS '支付方式',
    m.market_name AS '项目名称',
    o.hst_business_id AS '商家编码(合生通)',
    IFNULL(b.store_name,'') AS '商家名称(合生通)',
    o.shop_berth AS '铺位号',
    CASE o.pay_type
    WHEN 1 THEN '预缴费'
    WHEN 0 THEN '账单缴费'
    END AS '缴费类型',
    CASE o.btob_shop_no WHEN '651312105270019' THEN '上海大展投资管理有限公司'
                        WHEN '651312105270016' THEN '中先国际控股有限公司' END AS '公司名称'
    FROM hft_business_pay.bill_order o
    LEFT JOIN merchant_entity.entity_store b ON o.hst_business_id=b.store_id
    LEFT JOIN merchant_entity.market m ON o.market_id=m.id
    WHERE o.complete_date>=%(date_op)s
    AND o.complete_date<=concat(%(date_ed)s,' 23:59:59')
    AND o.market_id=%(market_id)s
    AND o.state=3"""
    billing_st_digital = """SELECT 
	m.`market_name`    AS '项目名称',
	b.`invoiceCode`    AS '缴款单号',
	b.`beginDT`        AS '计费日期',
	b.`paymentDT`      AS '最晚付款日期',
	b.`nTtotalAmount`  AS '不含税金额',
	b.`totalAmount`    AS '含税金额',
	b.`hstBusinessId`  AS '商铺编码（合生通）',
	(SELECT store_name 
	 FROM `merchant_entity`.`entity_store` s 
	   WHERE s.store_id = b.`hstBusinessId` )   AS '商家名称（合生通）',
	b.`businessCode`  AS '商铺编码',
	b.`brand`         AS '品牌名称',
	b.`shopBerth`     AS '铺位号',
	b.`shopName`      AS '单位名称',
	b.`payeeCompany`  AS '公司名称'
FROM shop_bill.bill b,merchant_entity.market m
WHERE  m.id = b.`marketId`
   AND b.`beginDT`>=%(date_op)s
   AND b.`beginDT`<=concat(%(date_ed)s,' 23:59:59')
   AND b.`marketId`=%(market_id)s"""
    billing_df_total   = pd.read_sql(billing_st_total,
                                     con=conn,
                                     params={'date_op': p_rq_start, 'date_ed': p_rq_end}
                                     )
    # 写【对账单】sheet页
    billing_wb = openpyxl.load_workbook('./billing_detail_110.xlsx')
    print('生成:【上海五角场合生汇项目】项目数字账单手续费数据...')
    billing_st_dzd = billing_wb['对账单']
    r=3
    for index,row in billing_df_total.iterrows():
        billing_st_dzd.cell(r, 1).value = row['项目名称']
        billing_st_dzd.cell(r, 2).value = row['时间维度']
        billing_st_dzd.cell(r, 3).value = row['总费用']
        billing_st_dzd.cell(r, 4).value = row['发布账单数']
        billing_st_dzd.cell(r, 5).value = row['电子签章费用']
        billing_st_dzd.cell(r, 6).value = row['缴费总金额']
        billing_st_dzd.cell(r, 7).value = row['缴费总手续费']
        billing_st_dzd.cell(r, 8).value = row['B2B缴费金额']
        billing_st_dzd.cell(r, 9).value = row['B2B缴费笔数']
        billing_st_dzd.cell(r, 10).value = row['B2B缴费手续费']
        billing_st_dzd.cell(r, 11).value = row['B2C缴费金额']
        billing_st_dzd.cell(r, 12).value = row['B2C缴费手续费']
        billing_st_dzd.cell(r, 13).value = row['扫码付缴费金额']
        billing_st_dzd.cell(r, 14).value = row['扫码付缴费手续费']
        r = r + 1
    # 写【缴费明细】sheet页
    billing_st_jfmx = billing_wb['缴费明细']
    billing_df_detail = pd.read_sql(billing_st_detail,con=conn,params={'market_id':110,'date_op': p_rq_start, 'date_ed': p_rq_end})
    r=2
    for index2, row2 in billing_df_detail.iterrows():
        billing_st_jfmx.cell(r,1).value = row2['交易流水号']
        billing_st_jfmx.cell(r,2).value = row2['缴费时间']
        billing_st_jfmx.cell(r,3).value = row2['缴费金额']
        billing_st_jfmx.cell(r,4).value = row2['手续费']
        billing_st_jfmx.cell(r,5).value = row2['支付方式']
        billing_st_jfmx.cell(r,6).value = row2['项目名称']
        billing_st_jfmx.cell(r,7).value = row2['商家编码(合生通)']
        billing_st_jfmx.cell(r,8).value = row2['商家名称(合生通)']
        billing_st_jfmx.cell(r,9).value = row2['铺位号']
        billing_st_jfmx.cell(r,10).value = row2['缴费类型']
        billing_st_jfmx.cell(r, 11).value = row2['公司名称']
        r = r + 1
    # 写【电子签章明细】sheet页
    billing_st_dzqz = billing_wb['电子签章明细']
    billing_df_digital = pd.read_sql(billing_st_digital, con=conn,params={'market_id': 110, 'date_op': p_rq_start, 'date_ed': p_rq_end})
    r = 2
    for index2, row2 in billing_df_digital.iterrows():
        billing_st_dzqz.cell(r, 1).value = row2['项目名称']
        billing_st_dzqz.cell(r, 2).value = row2['缴款单号']
        billing_st_dzqz.cell(r, 3).value = row2['计费日期']
        billing_st_dzqz.cell(r, 4).value = row2['最晚付款日期']
        billing_st_dzqz.cell(r, 5).value = row2['不含税金额']
        billing_st_dzqz.cell(r, 6).value = row2['含税金额']
        billing_st_dzqz.cell(r, 7).value = row2['商铺编码（合生通）']
        billing_st_dzqz.cell(r, 8).value = row2['商家名称（合生通）']
        billing_st_dzqz.cell(r, 9).value = row2['商铺编码']
        billing_st_dzqz.cell(r, 10).value = row2['品牌名称']
        billing_st_dzqz.cell(r, 11).value = row2['铺位号']
        billing_st_dzqz.cell(r, 12).value = row2['单位名称']
        billing_st_dzqz.cell(r, 13).value = row2['公司名称']
        r = r + 1
    billing_wb.save('./out/2024年上海五角场合生汇项目1季度数字账单手续费.xlsx')


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    output_dir = './out/'
    start_rq = '2024-01-01'
    end_rq = '2024-03-31'
    print('合生通项目组-数字账单手续费项目清单:')
    print('-----------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('-----------------------------')
    write_billing_total(start_rq,end_rq)