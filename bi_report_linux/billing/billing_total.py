#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os
import pymysql as py
import pandas as pd
import openpyxl
import warnings


def write_billing_total(p_rq_start,p_rq_end):
    conn = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',passwd='8Loxk2IJxaenJkE3', database='shop_bill')
    billing_st = """SELECT 
    	m.id AS '项目ID',
    	m.market_name AS '项目名称',
    	CONCAT(%(date_op)s,'~',%(date_ed)s) AS '时间维度',
    	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0)+ IFNULL(SUM(b.`电子签章费用`),0)AS '总费用',
        IFNULL(SUM(b.`本月发布账单数`),0) AS '发布账单数',
    	IFNULL(SUM(b.`电子签章费用`),0) AS '电子签章费用',
    	IFNULL(SUM(a.`B2B缴费金额`+a.`B2C缴费金额`+a.`扫码付缴费金额`),0) AS '缴费总金额',
    	IFNULL(SUM(a.`B2B缴费笔数`*5+a.`B2C缴费手续费`+a.`扫码付缴费手续费`),0) AS '缴费总手续费',   
    	IFNULL(SUM(a.`B2B缴费金额`),0) AS 'B2B缴费金额',
    	IFNULL(SUM(a.`B2B缴费笔数`),0) AS 'B2B缴费笔数',
    	IFNULL(SUM(a.`B2B缴费笔数`*5),0) AS 'B2B缴费手续费',
    	IFNULL(SUM(a.`B2C缴费金额`),0) AS 'B2C缴费金额',
    	IFNULL(SUM(a.`B2C缴费手续费`),0) AS 'B2C缴费手续费',
    	IFNULL(SUM(a.`扫码付缴费金额`),0) AS '扫码付缴费金额',
    	IFNULL(SUM(a.`扫码付缴费手续费`),0) AS '扫码付缴费手续费'
    FROM merchant_entity.market m 
     LEFT JOIN 
        (SELECT o.market_id,
    	ROUND(SUM(CASE WHEN o.tran_type='0002' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2B缴费金额',
    	SUM(CASE WHEN o.tran_type='0002' THEN 1 ELSE 0 END) AS 'B2B缴费笔数',
    	ROUND(SUM(CASE WHEN o.tran_type='0001' THEN o.total_fee ELSE 0 END)/100,2) AS 'B2C缴费金额',
    	SUM(CASE WHEN o.tran_type='0001' THEN ROUND(o.total_fee*0.002/100,2) ELSE 0 END) AS 'B2C缴费手续费',
    	ROUND(SUM(CASE WHEN o.tran_type='0009' THEN o.total_fee ELSE 0 END)/100,2) AS '扫码付缴费金额',
    	SUM(CASE WHEN o.tran_type='0009' THEN o.handling_fee/100 ELSE 0 END) AS '扫码付缴费手续费'
    FROM hft_business_pay.bill_order o
    WHERE o.complete_date>=%(date_op)s
    	AND o.complete_date<=concat(%(date_ed)s,' 23:59:59')
    	AND o.state=3	
    GROUP BY o.market_id) a  ON  m.id = a.market_id
    LEFT JOIN (
    SELECT b.`marketId`,
           COUNT(0) AS '本月发布账单数',
           COUNT(0)*0.4 AS '电子签章费用'
    FROM bill b
    WHERE  b.`beginDT`>=%(date_op)s
       AND b.`beginDT`<=concat(%(date_ed)s,' 23:59:59')
     GROUP BY b.`marketId`
    ) b ON m.id=b.`marketId`
    WHERE m.id IN(218,213,164,108,110,323,287,234,132,170,278,306,237,315,327,
                  150,155,159,183,188,194,203,249,252,277,1101,2181)
    GROUP BY m.`id`
    ORDER BY m.market_type,
    CASE 
    WHEN m.id =218 THEN 1
    WHEN m.id =213 THEN 2
    WHEN m.id =164 THEN 3
    WHEN m.id =108 THEN 4
    WHEN m.id =323 THEN 5
    WHEN m.id =287 THEN 6
    WHEN m.id =110 THEN 7
    WHEN m.id =234 THEN 8
    WHEN m.id =132 THEN 9
    WHEN m.id =170 THEN 10
    WHEN m.id =278 THEN 11
    WHEN m.id =306 THEN 12
    WHEN m.id =237 THEN 13
    WHEN m.id =315 THEN 14
    WHEN m.id =327 THEN 15
    WHEN m.id =150 THEN 16
    WHEN m.id =155 THEN 17
    WHEN m.id =159 THEN 18
    WHEN m.id =183 THEN 19
    WHEN m.id =188 THEN 20
    WHEN m.id =194 THEN 21
    WHEN m.id =203 THEN 22
    WHEN m.id =249 THEN 23
    WHEN m.id =252 THEN 24
    WHEN m.id =277 THEN 25
    WHEN m.id =1101 THEN 26
    WHEN m.id =2181 THEN 27
    ELSE 999 END"""
    billing_df = pd.read_sql(billing_st, con=conn, params={'date_op': p_rq_start, 'date_ed': p_rq_end},index_col='项目ID')
    billing_wb = openpyxl.load_workbook('./billing_total.xlsx')
    billing_st = billing_wb['数字账单手续费项目清单']
    r = 4
    for _,row in billing_df.iterrows():
        billing_st.cell(r, 3).value = row['项目名称']
        billing_st.cell(r, 4).value = row['时间维度']
        billing_st.cell(r, 5).value = row['总费用']
        billing_st.cell(r, 6).value = row['发布账单数']
        billing_st.cell(r, 7).value = row['电子签章费用']
        billing_st.cell(r, 8).value = row['缴费总金额']
        billing_st.cell(r, 9).value = row['缴费总手续费']
        billing_st.cell(r, 10).value = row['B2B缴费金额']
        billing_st.cell(r, 11).value = row['B2B缴费笔数']
        billing_st.cell(r, 12).value = row['B2B缴费手续费']
        billing_st.cell(r, 13).value = row['B2C缴费金额']
        billing_st.cell(r, 14).value = row['B2C缴费手续费']
        billing_st.cell(r, 15).value = row['扫码付缴费金额']
        billing_st.cell(r, 16).value = row['扫码付缴费手续费']
        r = r + 1
    billing_wb.save('./out/2024年4-6月数字账单手续费项目清单.xlsx')

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    output_dir = './out/'
    os.system('rm -rf ./out/*')
    start_rq = '2024-04-01'
    end_rq = '2024-06-30'
    print('合生通项目组-数字账单手续费项目清单:')
    print('-----------------------------')
    print('开始日期：', start_rq)
    print('结束日期：', end_rq)
    print('-----------------------------')
    write_billing_total(start_rq,end_rq)