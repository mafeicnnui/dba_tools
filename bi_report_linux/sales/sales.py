#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2023/12/29 9:06
# @Author : ma.fei
# @File : week_pos.py.py
# @Software: PyCharm

import os
import pymysql as py
import pandas as pd
import openpyxl
import warnings

def write_market_sales_data():
    conn = py.connect(host='rm-2zer0v9g25bgu4rx43o.mysql.rds.aliyuncs.com', user='hopsonone_do',passwd='8Loxk2IJxaenJkE3', database='shop_bill')
    market_names = """SELECT id AS market_id,market_name 
                            FROM `merchant_entity`.market 
                            WHERE id IN(218,110,108,132,164,213,237,278,287,306,327,315)"""
    sales_amount = """SELECT b.m_mobile,      
       ROUND(SUM((CASE WHEN d.state=0 THEN d.consume_amount WHEN d.state IN(1,4,5) THEN -d.consume_amount ELSE 0 END))/100,2) AS consume_amount,
       e.store_name
FROM `hopsonone_point`.`members_points_detail` d,
     `hopsonone_members`.`members` b,
     `merchant_entity`.`entity_store` e
WHERE d.m_id=b.`m_id`
  AND d.`business_id`=e.store_id
  AND d.market_id=%(market_id)s
  AND d.consume_time >= %(date_op)s 
  and d.consume_time <= concat(%(date_ed)s,' 23:59:59')
GROUP BY b.m_mobile,e.store_name
ORDER BY 2 DESC LIMIT 500"""
    sales_point = """SELECT b.m_mobile,      
           IFNULL(SUM(CASE WHEN state=0 THEN actual_point WHEN state IN(1,4,5) THEN -actual_point ELSE 0 END),0) AS actual_point,
           e.store_name
    FROM `hopsonone_point`.`members_points_detail` d,
         `hopsonone_members`.`members` b,
         `merchant_entity`.`entity_store` e
    WHERE d.m_id=b.`m_id`
      AND d.`business_id`=e.store_id
      AND d.market_id=%(market_id)s
      AND d.consume
_time >= %(date_op)s 
      and d.consume_time <= concat(%(date_ed)s,' 23:59:59')
    GROUP BY b.m_mobile,e.store_name
    ORDER BY 2 DESC LIMIT 500"""

    df_market_names = pd.read_sql(market_names,con=conn,index_col='market_id')
    for index,row in df_market_names.iterrows():
        print('正在生成{}项目消费&积分数据...'.format(row['market_name']))
        sales_wb = openpyxl.load_workbook('./sales.xlsx')
        # 写【2023年全年会员消费TOP100】sheet页
        wb_sales_amount_2023 = sales_wb['2023年全年会员消费TOP100']
        df_sales_amount_2023 = pd.read_sql(
                                    sales_amount,con=conn,
                                    params={'market_id':index,'date_op': '2023-01-01','date_ed': '2023-12-31'})
        r=2
        for index2, row2 in df_sales_amount_2023.iterrows():
            wb_sales_amount_2023.cell(r,1).value = row2['m_mobile']
            wb_sales_amount_2023.cell(r,2).value = row2['consume_amount']
            wb_sales_amount_2023.cell(r,3).value = row2['store_name']
            r = r + 1
        # 写【2024年1-8月年会员消费TOP100】sheet页
        wb_sales_amount_2024 = sales_wb['2024年1-8月会员消费TOP100']
        df_sales_amount_2024 = pd.read_sql(
                                  sales_amount, con=conn,
                                  params={'market_id': index, 'date_op': '2024-01-01','date_ed': '2024-08-31'})
        r = 2
        for index2, row2 in df_sales_amount_2024.iterrows():
            wb_sales_amount_2024.cell(r, 1).value = row2['m_mobile']
            wb_sales_amount_2024.cell(r, 2).value = row2['consume_amount']
            wb_sales_amount_2024.cell(r, 3).value = row2['store_name']
            r = r + 1

        # 写【2023年全年会员销费积分TOP100】sheet页
        wb_sales_point_2023 = sales_wb['2023年店铺会员消费积分TOP100']
        df_sales_point_2023 = pd.read_sql(
                    sales_point, con=conn,
                    params={'market_id': index, 'date_op': '2023-01-01', 'date_ed': '2023-12-31'})
        r = 2
        for index2, row2 in df_sales_point_2023.iterrows():
            wb_sales_point_2023.cell(r, 1).value = row2['m_mobile']
            wb_sales_point_2023.cell(r, 2).value = row2['actual_point']
            wb_sales_point_2023.cell(r, 3).value = row2['store_name']
            r = r + 1

        # 写【2024年1-8月年会员消费积分TOP100】sheet页
        wb_sales_point_2024 = sales_wb['2024年店铺会员消费积分TOP100']
        df_sales_point_2024 = pd.read_sql(
                        sales_point, con=conn,
                        params={'market_id': index, 'date_op': '2024-01-01', 'date_ed': '2024-08-31'})
        r = 2
        for index2, row2 in df_sales_point_2024.iterrows():
            wb_sales_point_2024.cell(r, 1).value = row2['m_mobile']
            wb_sales_point_2024.cell(r, 2).value = row2['actual_point']
            wb_sales_point_2024.cell(r, 3).value = row2['store_name']
            r = r + 1
        sales_wb.save('./out/'+row['market_name']+'-2023~2024消费&积分Top100.xlsx')

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    write_market_sales_data()